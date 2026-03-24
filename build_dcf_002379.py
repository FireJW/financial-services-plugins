"""
宏桥控股 (002379.SZ) DCF Valuation Model
中国宏桥集团核心铝业资产 A 股平台
Generated: 2026-03-09
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()

# ============================================================
# Color & Style Definitions
# ============================================================
BLUE_FONT = Font(name="Microsoft YaHei", size=11, color="0000FF")       # Hardcoded inputs
BLACK_FONT = Font(name="Microsoft YaHei", size=11, color="000000")      # Formulas
GREEN_FONT = Font(name="Microsoft YaHei", size=11, color="008000")      # Cross-sheet links
HEADER_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="000000")
TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color="000000")
RESULT_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="000000")

DARK_BLUE_FILL = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
THICK_BOTTOM = Border(bottom=Side(style="medium"))

PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
NUM_FMT_1 = '#,##0.0'
PRICE_FMT = '¥#,##0.00'
RATIO_FMT = '0.00'

def set_cell(ws, row, col, value, font=None, fill=None, fmt=None, align=None, comment_text=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    if align: cell.alignment = align
    cell.border = THIN_BORDER
    if comment_text:
        cell.comment = openpyxl.comments.Comment(comment_text, "DCF Model")
    return cell

def section_header(ws, row, text, col_start=1, col_end=10):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = DARK_BLUE_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.cell(row=row, column=col_start, value=text)
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)

def sub_header(ws, row, text, col_start=1, col_end=10):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = LIGHT_BLUE_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.cell(row=row, column=col_start, value=text)

# ============================================================
# HISTORICAL DATA (Pro-forma with 宏拓实业 assets)
# All figures in ¥ Millions (百万元)
# ============================================================
COMPANY = "宏桥控股"
TICKER = "002379.SZ"
DATE = "2026-03-09"

# Market Data
STOCK_PRICE = 30.29        # ¥ per share (2026-03-02)
SHARES_OUT = 12570         # Million shares (total ~125.7亿, post-restructuring)
MARKET_CAP = STOCK_PRICE * SHARES_OUT  # ~380,769 ¥M

# Historical P&L (Pro-forma, China Hongqiao consolidated)
hist = {
    "year":     [2021,    2022,     2023,     2024],
    "revenue":  [114500,  131700,   133620,   156169],   # ¥M
    "cogs":     [91600,   113400,   112630,   114000],   # Cost of sales (estimated)
    "gp":       [22900,   18300,    20990,    42169],    # Gross profit
    "gp_pct":   [0.200,   0.139,    0.157,    0.270],    # Gross margin
    "sga":      [2200,    2500,     2600,     2800],     # SG&A
    "rd":       [800,     900,      950,      1000],     # R&D
    "ebit":     [19900,   14900,    17440,    38369],    # Operating profit
    "ebit_pct": [0.174,   0.113,    0.130,    0.246],    # EBIT margin
    "tax":      [3580,    2682,     3139,     6907],     # Income tax
    "ni":       [16030,   8702,     11460,    22372],    # Net income attr. to parent
    "da":       [8000,    8500,     9000,     9500],     # D&A (estimated)
    "capex":    [4500,    5000,     5010,     5000],     # CapEx
    "ebitda":   [27900,   23400,    26440,    47869],    # EBITDA
}

# Balance Sheet Items (2024)
TOTAL_DEBT = 45000    # ¥M interest-bearing debt (estimated from net gearing ~23%)
CASH = 28000          # ¥M cash & equivalents
NET_DEBT = TOTAL_DEBT - CASH  # 17,000 ¥M

# WACC Inputs
RF_RATE = 0.022        # China 10Y gov bond ~2.2%
BETA = 1.14            # From investing.com
ERP = 0.065            # A-share equity risk premium
COST_OF_EQUITY = RF_RATE + BETA * ERP  # ~9.61%
PRE_TAX_DEBT = 0.038   # ~3.8% (LPR-based)
TAX_RATE = 0.18        # Effective tax rate (~18% due to preferential policies)
AFTER_TAX_DEBT = PRE_TAX_DEBT * (1 - TAX_RATE)
EV_CALC = MARKET_CAP + NET_DEBT
EQ_WEIGHT = MARKET_CAP / EV_CALC
DEBT_WEIGHT = NET_DEBT / EV_CALC
WACC = COST_OF_EQUITY * EQ_WEIGHT + AFTER_TAX_DEBT * DEBT_WEIGHT

# Projection Assumptions (5-year, Base Case)
proj = {
    "years": [2025, 2026, 2027, 2028, 2029],
    "bear":  {"rev_growth": [0.04, 0.03, 0.02, 0.02, 0.02],
              "ebit_margin": [0.22, 0.21, 0.20, 0.19, 0.18]},
    "base":  {"rev_growth": [0.08, 0.06, 0.05, 0.04, 0.03],
              "ebit_margin": [0.24, 0.23, 0.22, 0.21, 0.20]},
    "bull":  {"rev_growth": [0.12, 0.10, 0.08, 0.06, 0.05],
              "ebit_margin": [0.27, 0.26, 0.25, 0.24, 0.23]},
}
DA_PCT = 0.060          # D&A as % of revenue
CAPEX_PCT = 0.035       # CapEx as % of revenue
NWC_PCT = 0.01          # NWC change as % of revenue change
TERM_GROWTH = {"bear": 0.025, "base": 0.03, "bull": 0.04}

# ============================================================
# DCF SHEET
# ============================================================
ws = wb.active
ws.title = "DCF"
ws.sheet_properties.tabColor = "17365D"

# Column widths
ws.column_dimensions['A'].width = 32
for c in range(2, 12):
    ws.column_dimensions[get_column_letter(c)].width = 16

# --- Row 1-3: Title ---
R = 1
set_cell(ws, R, 1, f"{COMPANY} ({TICKER}) DCF 估值模型", TITLE_FONT)
ws.merge_cells('A1:J1')
R = 2
set_cell(ws, R, 1, f"代码: {TICKER} | 日期: {DATE} | 报告期: FY2024", BLACK_FONT)
ws.merge_cells('A2:J2')
R = 3
set_cell(ws, R, 1, "单位: 人民币百万元 (¥M)，每股数据除外 | Unit: RMB Millions except per-share", BLACK_FONT)
ws.merge_cells('A3:J3')

# --- Row 5-6: Case Selector ---
R = 5
section_header(ws, R, "情景选择 / CASE SELECTOR", 1, 10)
R = 6
set_cell(ws, R, 1, "当前情景 / Active Case (1=悲观Bear, 2=基准Base, 3=乐观Bull)", SUBHEADER_FONT, YELLOW_FILL)
set_cell(ws, R, 2, 2, BLUE_FONT, YELLOW_FILL, align=CENTER,
         comment_text="Source: User input. Change to 1 (Bear), 2 (Base), or 3 (Bull)")

# --- Row 8: Market Data ---
R = 8
section_header(ws, R, "市场数据 / MARKET DATA (非情景相关 / Not case dependent)", 1, 10)
R = 9
labels_market = [
    ("当前股价 / Stock Price (¥)", STOCK_PRICE, PRICE_FMT, "Source: 东方财富 2026-03-02 收盘价"),
    ("总股本 / Shares Outstanding (M)", SHARES_OUT, NUM_FMT, "Source: 深交所 2026-01-13 重组后新增股份上市"),
    ("总市值 / Market Cap (¥M)", None, NUM_FMT, None),  # formula
    ("有息负债 / Total Debt (¥M)", TOTAL_DEBT, NUM_FMT, "Source: 中国宏桥 2024年报 估计值"),
    ("现金 / Cash (¥M)", CASH, NUM_FMT, "Source: 中国宏桥 2024年报 估计值"),
    ("净负债 / Net Debt (¥M)", None, NUM_FMT, None),  # formula
    ("企业价值 / Enterprise Value (¥M)", None, NUM_FMT, None),  # formula
]
for i, (label, val, fmt, cmt) in enumerate(labels_market):
    r = R + i
    set_cell(ws, r, 1, label, SUBHEADER_FONT, LIGHT_GRAY_FILL)
    if val is not None:
        set_cell(ws, r, 2, val, BLUE_FONT, LIGHT_GREEN_FILL, fmt, CENTER, cmt)
    elif i == 2:  # Market Cap = Price * Shares
        set_cell(ws, r, 2, f"=B9*B10", BLACK_FONT, fmt=fmt, align=CENTER)
    elif i == 5:  # Net Debt = Debt - Cash
        set_cell(ws, r, 2, f"=B12-B13", BLACK_FONT, fmt=fmt, align=CENTER)
    elif i == 6:  # EV = MCap + Net Debt
        set_cell(ws, r, 2, f"=B11+B14", BLACK_FONT, fmt=fmt, align=CENTER)

# --- Row 17: Scenario Assumptions ---
R = 17
section_header(ws, R, "情景假设 / SCENARIO ASSUMPTIONS", 1, 10)

# Bear Case Block
R = 18
sub_header(ws, R, "悲观情景 / BEAR CASE", 1, 10)
R = 19
headers_proj = ["假设 / Assumption"] + [f"FY{y}E" for y in proj["years"]]
for c, h in enumerate(headers_proj, 1):
    set_cell(ws, R, c, h, SUBHEADER_FONT, LIGHT_BLUE_FILL, align=CENTER)

bear_rows = [
    ("营收增速 / Revenue Growth", proj["bear"]["rev_growth"], PCT_FMT),
    ("EBIT利润率 / EBIT Margin", proj["bear"]["ebit_margin"], PCT_FMT),
    ("永续增长率 / Terminal Growth", [TERM_GROWTH["bear"]], PCT_FMT),
]
R = 20
for label, vals, fmt in bear_rows:
    set_cell(ws, R, 1, label, SUBHEADER_FONT, LIGHT_GRAY_FILL)
    for j, v in enumerate(vals):
        set_cell(ws, R, 2+j, v, BLUE_FONT, LIGHT_GREEN_FILL, fmt, CENTER)
    R += 1

# Base Case Block
R = 23
sub_header(ws, R, "基准情景 / BASE CASE", 1, 10)
R = 24
for c, h in enumerate(headers_proj, 1):
    set_cell(ws, R, c, h, SUBHEADER_FONT, LIGHT_BLUE_FILL, align=CENTER)
R = 25
base_rows = [
    ("营收增速 / Revenue Growth", proj["base"]["rev_growth"], PCT_FMT),
    ("EBIT利润率 / EBIT Margin", proj["base"]["ebit_margin"], PCT_FMT),
    ("永续增长率 / Terminal Growth", [TERM_GROWTH["base"]], PCT_FMT),
]
for label, vals, fmt in base_rows:
    set_cell(ws, R, 1, label, SUBHEADER_FONT, LIGHT_GRAY_FILL)
    for j, v in enumerate(vals):
        set_cell(ws, R, 2+j, v, BLUE_FONT, LIGHT_GREEN_FILL, fmt, CENTER)
    R += 1

# Bull Case Block
R = 28
sub_header(ws, R, "乐观情景 / BULL CASE", 1, 10)
R = 29
for c, h in enumerate(headers_proj, 1):
    set_cell(ws, R, c, h, SUBHEADER_FONT, LIGHT_BLUE_FILL, align=CENTER)
R = 30
bull_rows = [
    ("营收增速 / Revenue Growth", proj["bull"]["rev_growth"], PCT_FMT),
    ("EBIT利润率 / EBIT Margin", proj["bull"]["ebit_margin"], PCT_FMT),
    ("永续增长率 / Terminal Growth", [TERM_GROWTH["bull"]], PCT_FMT),
]
for label, vals, fmt in bull_rows:
    set_cell(ws, R, 1, label, SUBHEADER_FONT, LIGHT_GRAY_FILL)
    for j, v in enumerate(vals):
        set_cell(ws, R, 2+j, v, BLUE_FONT, LIGHT_GREEN_FILL, fmt, CENTER)
    R += 1

# --- Consolidation Row (Selected Case via INDEX) ---
R = 33
sub_header(ws, R, "选定情景 / SELECTED CASE (由情景选择器驱动 / Driven by Case Selector)", 1, 10)
R = 34
for c, h in enumerate(headers_proj, 1):
    set_cell(ws, R, c, h, SUBHEADER_FONT, LIGHT_BLUE_FILL, align=CENTER)

# Row 35: Selected Revenue Growth
R = 35
set_cell(ws, R, 1, "营收增速 / Revenue Growth", SUBHEADER_FONT, YELLOW_FILL)
for j in range(5):
    col_letter = get_column_letter(2+j)
    # INDEX picks from Bear(row20), Base(row25), Bull(row30) based on B6
    bear_cell = f"${col_letter}$20"
    base_cell = f"${col_letter}$25"
    bull_cell = f"${col_letter}$30"
    formula = f'=IF($B$6=1,{bear_cell},IF($B$6=2,{base_cell},{bull_cell}))'
    set_cell(ws, R, 2+j, formula, BLACK_FONT, YELLOW_FILL, PCT_FMT, CENTER)

# Row 36: Selected EBIT Margin
R = 36
set_cell(ws, R, 1, "EBIT利润率 / EBIT Margin", SUBHEADER_FONT, YELLOW_FILL)
for j in range(5):
    col_letter = get_column_letter(2+j)
    bear_cell = f"${col_letter}$21"
    base_cell = f"${col_letter}$26"
    bull_cell = f"${col_letter}$31"
    formula = f'=IF($B$6=1,{bear_cell},IF($B$6=2,{base_cell},{bull_cell}))'
    set_cell(ws, R, 2+j, formula, BLACK_FONT, YELLOW_FILL, PCT_FMT, CENTER)

# Row 37: Selected Terminal Growth
R = 37
set_cell(ws, R, 1, "永续增长率 / Terminal Growth", SUBHEADER_FONT, YELLOW_FILL)
bear_cell = "$B$22"
base_cell = "$B$27"
bull_cell = "$B$32"
formula = f'=IF($B$6=1,{bear_cell},IF($B$6=2,{base_cell},{bull_cell}))'
set_cell(ws, R, 2, formula, BLACK_FONT, YELLOW_FILL, PCT_FMT, CENTER)

# Row 38: Common assumptions
R = 38
set_cell(ws, R, 1, "折旧摊销占收入% / D&A % Rev", SUBHEADER_FONT, LIGHT_GRAY_FILL)
set_cell(ws, R, 2, DA_PCT, BLUE_FONT, LIGHT_GREEN_FILL, PCT_FMT, CENTER,
         "Source: 历史均值 ~6% / Historical average")
R = 39
set_cell(ws, R, 1, "资本开支占收入% / CapEx % Rev", SUBHEADER_FONT, LIGHT_GRAY_FILL)
set_cell(ws, R, 2, CAPEX_PCT, BLUE_FONT, LIGHT_GREEN_FILL, PCT_FMT, CENTER,
         "Source: 2023A CapEx 50.1亿 / Rev 1336亿 ≈ 3.5%")
R = 40
set_cell(ws, R, 1, "营运资本变动占ΔRev% / NWC % ΔRev", SUBHEADER_FONT, LIGHT_GRAY_FILL)
set_cell(ws, R, 2, NWC_PCT, BLUE_FONT, LIGHT_GREEN_FILL, PCT_FMT, CENTER,
         "Source: 估计值 1% / Estimated")
R = 41
set_cell(ws, R, 1, "有效税率 / Tax Rate", SUBHEADER_FONT, LIGHT_GRAY_FILL)
set_cell(ws, R, 2, TAX_RATE, BLUE_FONT, LIGHT_GREEN_FILL, PCT_FMT, CENTER,
         "Source: 中国宏桥有效税率 ~18% (含优惠政策)")

# ============================================================
# Row 43+: HISTORICAL & PROJECTED FINANCIALS
# ============================================================
R = 43
section_header(ws, R, "利润表与现金流预测 / INCOME STATEMENT & CASH FLOW PROJECTION", 1, 10)

# Column Layout: A=Label, B-E=Historical (2021-2024), F-J=Projected (2025E-2029E)
R = 44
all_years = [""] + [f"FY{y}A" for y in hist["year"]] + [f"FY{y}E" for y in proj["years"]]
for c, h in enumerate(all_years, 1):
    fill = LIGHT_BLUE_FILL if "A" in str(h) else LIGHT_GRAY_FILL if "E" in str(h) else None
    set_cell(ws, R, c, h, SUBHEADER_FONT, fill, align=CENTER)

# Revenue row (45)
R = 45
set_cell(ws, R, 1, "营业收入 / Revenue", SUBHEADER_FONT)
for i, rev in enumerate(hist["revenue"]):
    set_cell(ws, R, 2+i, rev, BLUE_FONT, fmt=NUM_FMT, align=CENTER,
             comment_text=f"Source: 中国宏桥 FY{hist['year'][i]}年报")
# Projected revenue: =prior*(1+growth)
for j in range(5):
    prior_col = get_column_letter(5+j)  # E=2024, F=2025E, etc.
    growth_col = get_column_letter(2+j)  # growth from row 35
    formula = f"={prior_col}45*(1+{growth_col}$35)"
    set_cell(ws, R, 6+j, formula, BLACK_FONT, fmt=NUM_FMT, align=CENTER)

# Revenue Growth row (46)
R = 46
set_cell(ws, R, 1, "  增速 / Growth %", BLACK_FONT)
for i in range(1, len(hist["revenue"])):
    prev_col = get_column_letter(1+i)
    cur_col = get_column_letter(2+i)
    formula = f"={cur_col}45/{prev_col}45-1"
    set_cell(ws, R, 2+i, formula, BLACK_FONT, fmt=PCT_FMT, align=CENTER)
for j in range(5):
    col = get_column_letter(6+j)
    prev = get_column_letter(5+j)
    formula = f"={col}45/{prev}45-1"
    set_cell(ws, R, 6+j, formula, BLACK_FONT, fmt=PCT_FMT, align=CENTER)

# EBIT row (47)
R = 47
set_cell(ws, R, 1, "营业利润 / EBIT", SUBHEADER_FONT)
for i, ebit in enumerate(hist["ebit"]):
    set_cell(ws, R, 2+i, ebit, BLUE_FONT, fmt=NUM_FMT, align=CENTER)
for j in range(5):
    rev_col = get_column_letter(6+j)
    margin_col = get_column_letter(2+j)
    formula = f"={rev_col}45*{margin_col}$36"
    set_cell(ws, R, 6+j, formula, BLACK_FONT, fmt=NUM_FMT, align=CENTER)

# EBIT Margin row (48)
R = 48
set_cell(ws, R, 1, "  EBIT利润率 / EBIT Margin", BLACK_FONT)
for i in range(len(hist["year"])):
    col = get_column_letter(2+i)
    formula = f"={col}47/{col}45"
    set_cell(ws, R, 2+i, formula, BLACK_FONT, fmt=PCT_FMT, align=CENTER)
for j in range(5):
    col = get_column_letter(6+j)
    formula = f"={col}47/{col}45"
    set_cell(ws, R, 6+j, formula, BLACK_FONT, fmt=PCT_FMT, align=CENTER)

# Tax row (49)
R = 49
set_cell(ws, R, 1, "(-) 所得税 / Taxes", BLACK_FONT)
for i, tax in enumerate(hist["tax"]):
    set_cell(ws, R, 2+i, -abs(tax), BLUE_FONT, fmt=NUM_FMT, align=CENTER)
for j in range(5):
    col = get_column_letter(6+j)
    formula = f"=-{col}47*$B$41"
    set_cell(ws, R, 6+j, formula, BLACK_FONT, fmt=NUM_FMT, align=CENTER)

# NOPAT row (50)
R = 50
set_cell(ws, R, 1, "NOPAT (税后净营业利润)", SUBHEADER_FONT)
for i in range(len(hist["year"])):
    col = get_column_letter(2+i)
    formula = f"={col}47+{col}49"
    set_cell(ws, R, 2+i, formula, BLACK_FONT, fmt=NUM_FMT, align=CENTER)
for j in range(5):
    col = get_column_letter(6+j)
    formula = f"={col}47+{col}49"
    set_cell(ws, R, 6+j, formula, BLACK_FONT, fmt=NUM_FMT, align=CENTER)

# D&A row (51)
R = 51
set_cell(ws, R, 1, "(+) 折旧摊销 / D&A", BLACK_FONT)
for i, da in enumerate(hist["da"]):
    set_cell(ws, R, 2+i, da, BLUE_FONT, fmt=NUM_FMT, align=CENTER)
for j in range(5):
    col = get_column_letter(6+j)
    formula = f"={col}45*$B$38"
    set_cell(ws, R, 6+j, formula, BLACK_FONT, fmt=NUM_FMT, align=CENTER)

# CapEx row (52)
R = 52
set_cell(ws, R, 1, "(-) 资本开支 / CapEx", BLACK_FONT)
for i, capex in enumerate(hist["capex"]):
    set_cell(ws, R, 2+i, -abs(capex), BLUE_FONT, fmt=NUM_FMT, align=CENTER)
for j in range(5):
    col = get_column_letter(6+j)
    formula = f"=-{col}45*$B$39"
    set_cell(ws, R, 6+j, formula, BLACK_FONT, fmt=NUM_FMT, align=CENTER)

# NWC row (53)
R = 53
set_cell(ws, R, 1, "(-) Δ营运资本 / Δ NWC", BLACK_FONT)
set_cell(ws, R, 2, 0, BLUE_FONT, fmt=NUM_FMT, align=CENTER)
for i in range(1, len(hist["year"])):
    set_cell(ws, R, 2+i, 0, BLUE_FONT, fmt=NUM_FMT, align=CENTER)
for j in range(5):
    col = get_column_letter(6+j)
    prev = get_column_letter(5+j)
    formula = f"=-({col}45-{prev}45)*$B$40"
    set_cell(ws, R, 6+j, formula, BLACK_FONT, fmt=NUM_FMT, align=CENTER)

# UFCF row (54)
R = 54
set_cell(ws, R, 1, "无杠杆自由现金流 / Unlevered FCF", SUBHEADER_FONT, LIGHT_GRAY_FILL)
for i in range(len(hist["year"])):
    col = get_column_letter(2+i)
    formula = f"={col}50+{col}51+{col}52+{col}53"
    set_cell(ws, R, 2+i, formula, RESULT_FONT, LIGHT_GRAY_FILL, NUM_FMT, CENTER)
for j in range(5):
    col = get_column_letter(6+j)
    formula = f"={col}50+{col}51+{col}52+{col}53"
    set_cell(ws, R, 6+j, formula, RESULT_FONT, LIGHT_GRAY_FILL, NUM_FMT, CENTER)

# ============================================================
# Row 56+: DCF VALUATION
# ============================================================
R = 56
section_header(ws, R, "DCF 估值 / DCF VALUATION", 1, 10)

# Discount factors
R = 57
set_cell(ws, R, 1, "WACC", SUBHEADER_FONT, YELLOW_FILL)
set_cell(ws, R, 2, f"=WACC!B16", GREEN_FONT, YELLOW_FILL, PCT_FMT, CENTER)

R = 58
set_cell(ws, R, 1, "折现期 / Discount Period", SUBHEADER_FONT)
for j in range(5):
    set_cell(ws, R, 6+j, 0.5 + j, BLUE_FONT, fmt=NUM_FMT_1, align=CENTER)

R = 59
set_cell(ws, R, 1, "折现因子 / Discount Factor", BLACK_FONT)
for j in range(5):
    col = get_column_letter(6+j)
    formula = f"=1/(1+$B$57)^{col}58"
    set_cell(ws, R, 6+j, formula, BLACK_FONT, fmt='0.0000', align=CENTER)

R = 60
set_cell(ws, R, 1, "FCF 现值 / PV of FCF", SUBHEADER_FONT)
for j in range(5):
    col = get_column_letter(6+j)
    formula = f"={col}54*{col}59"
    set_cell(ws, R, 6+j, formula, BLACK_FONT, fmt=NUM_FMT, align=CENTER)

# Terminal Value
R = 62
set_cell(ws, R, 1, "终值 FCF / Terminal FCF", BLACK_FONT)
set_cell(ws, R, 2, "=J54*(1+B37)", BLACK_FONT, fmt=NUM_FMT, align=CENTER)

R = 63
set_cell(ws, R, 1, "终值 / Terminal Value", BLACK_FONT)
set_cell(ws, R, 2, "=B62/(B57-B37)", BLACK_FONT, fmt=NUM_FMT, align=CENTER)

R = 64
set_cell(ws, R, 1, "终值现值 / PV of Terminal Value", SUBHEADER_FONT)
set_cell(ws, R, 2, "=B63/(1+B57)^4.5", BLACK_FONT, fmt=NUM_FMT, align=CENTER)

# Valuation Summary
R = 66
section_header(ws, R, "估值汇总 / VALUATION SUMMARY", 1, 10)

R = 67
set_cell(ws, R, 1, "预测期FCF现值合计 / Sum PV of FCFs", SUBHEADER_FONT, LIGHT_GRAY_FILL)
set_cell(ws, R, 2, "=SUM(F60:J60)", BLACK_FONT, LIGHT_GRAY_FILL, NUM_FMT, CENTER)

R = 68
set_cell(ws, R, 1, "终值现值 / PV of Terminal Value", SUBHEADER_FONT, LIGHT_GRAY_FILL)
set_cell(ws, R, 2, "=B64", BLACK_FONT, LIGHT_GRAY_FILL, NUM_FMT, CENTER)

R = 69
set_cell(ws, R, 1, "企业价值 / Enterprise Value", SUBHEADER_FONT, YELLOW_FILL)
set_cell(ws, R, 2, "=B67+B68", RESULT_FONT, YELLOW_FILL, NUM_FMT, CENTER)

R = 70
set_cell(ws, R, 1, "(-) 净负债 / Net Debt", SUBHEADER_FONT, LIGHT_GRAY_FILL)
set_cell(ws, R, 2, "=-B14", BLACK_FONT, LIGHT_GRAY_FILL, NUM_FMT, CENTER)

R = 71
set_cell(ws, R, 1, "股权价值 / Equity Value", SUBHEADER_FONT, YELLOW_FILL)
set_cell(ws, R, 2, "=B69+B70", RESULT_FONT, YELLOW_FILL, NUM_FMT, CENTER)

R = 72
set_cell(ws, R, 1, "总股本(M) / Shares Outstanding", BLACK_FONT)
set_cell(ws, R, 2, "=B10", GREEN_FONT, fmt=NUM_FMT, align=CENTER)

R = 73
set_cell(ws, R, 1, "每股内在价值 / IMPLIED PRICE (¥)", Font(name="Microsoft YaHei", size=14, bold=True, color="FF0000"), YELLOW_FILL)
set_cell(ws, R, 2, "=B71/B72", Font(name="Microsoft YaHei", size=14, bold=True, color="FF0000"), YELLOW_FILL, PRICE_FMT, CENTER)

R = 74
set_cell(ws, R, 1, "当前股价 / Current Price (¥)", SUBHEADER_FONT)
set_cell(ws, R, 2, "=B9", GREEN_FONT, fmt=PRICE_FMT, align=CENTER)

R = 75
set_cell(ws, R, 1, "隐含回报 / Implied Upside/(Downside)", SUBHEADER_FONT, YELLOW_FILL)
set_cell(ws, R, 2, "=B73/B74-1", RESULT_FONT, YELLOW_FILL, PCT_FMT, CENTER)

# ============================================================
# Row 77+: SENSITIVITY ANALYSIS
# ============================================================
R = 77
section_header(ws, R, "敏感性分析 / SENSITIVITY ANALYSIS", 1, 10)

# Table 1: WACC vs Terminal Growth
R = 78
sub_header(ws, R, "表1: 每股价值 vs WACC & 永续增长率 / Implied Price vs WACC & Terminal Growth", 1, 8)
R = 79
# Column headers: Terminal growth
tg_vals = [0.02, 0.025, 0.03, 0.035, 0.04]
wacc_vals = [0.07, 0.08, 0.09, 0.10, 0.11]

set_cell(ws, R, 1, "WACC \\ TGR", SUBHEADER_FONT, LIGHT_BLUE_FILL, align=CENTER)
for j, tg in enumerate(tg_vals):
    set_cell(ws, R, 2+j, tg, SUBHEADER_FONT, LIGHT_BLUE_FILL, PCT_FMT, CENTER)

for i, wacc_v in enumerate(wacc_vals):
    r = R + 1 + i
    set_cell(ws, r, 1, wacc_v, BLUE_FONT, fmt=PCT_FMT, align=CENTER)
    for j, tg_v in enumerate(tg_vals):
        # Full DCF recalc: (sum_pv_fcf_with_wacc + TV_pv_with_wacc_tg - net_debt) / shares
        # Simplified: use FCF from row 54 and recalculate discount + terminal
        # PV_FCFs = sum of FCF_year / (1+wacc)^period
        pv_parts = []
        for k in range(5):
            fcf_col = get_column_letter(6+k)
            period = 0.5 + k
            pv_parts.append(f"{fcf_col}54/(1+$A{r})^{period}")
        sum_pv = "+".join(pv_parts)
        
        # Terminal: J54*(1+tg)/(wacc-tg) / (1+wacc)^4.5
        tg_cell = f"${get_column_letter(2+j)}${R}"
        wacc_cell = f"$A{r}"
        tv = f"(J54*(1+{tg_cell})/({wacc_cell}-{tg_cell}))/(1+{wacc_cell})^4.5"
        
        formula = f"=({sum_pv}+{tv}+B70)/B72"
        set_cell(ws, r, 2+j, formula, BLACK_FONT, fmt=PRICE_FMT, align=CENTER)

# Table 2: Revenue Growth vs EBIT Margin (using Year 1 assumptions)
R = 86
sub_header(ws, R, "表2: 每股价值 vs 营收增速(Y1) & EBIT利润率(Y1) / Price vs Rev Growth & EBIT Margin", 1, 8)
R = 87
rg_vals = [0.04, 0.06, 0.08, 0.10, 0.12]
em_vals = [0.18, 0.20, 0.22, 0.24, 0.27]

set_cell(ws, R, 1, "RevGr \\ EBIT%", SUBHEADER_FONT, LIGHT_BLUE_FILL, align=CENTER)
for j, em in enumerate(em_vals):
    set_cell(ws, R, 2+j, em, SUBHEADER_FONT, LIGHT_BLUE_FILL, PCT_FMT, CENTER)

for i, rg in enumerate(rg_vals):
    r = R + 1 + i
    set_cell(ws, r, 1, rg, BLUE_FONT, fmt=PCT_FMT, align=CENTER)
    for j, em_v in enumerate(em_vals):
        # Quick Y1 FCF approximation: rev*(1+rg)*em*(1-tax) + rev*(1+rg)*da% - rev*(1+rg)*capex%
        # Then apply multiplier based on growth and discount
        # Using simplified perpetuity: FCF1 / (WACC - g_avg) for quick sensitivity
        em_cell = f"${get_column_letter(2+j)}${R}"
        rg_cell = f"$A{r}"
        # Y1 Revenue
        rev_base = "E45"  # 2024 actual
        # Y1 FCF = Rev*(1+rg)*ebit_margin*(1-tax) + Rev*(1+rg)*da% - Rev*(1+rg)*capex% - delta_rev*nwc%
        fcf1 = f"({rev_base}*(1+{rg_cell})*{em_cell}*(1-$B$41)+{rev_base}*(1+{rg_cell})*$B$38-{rev_base}*(1+{rg_cell})*$B$39-{rev_base}*{rg_cell}*$B$40)"
        # Simple 5-year perpetuity value: FCF1 * (1 - (1+g)^5/(1+wacc)^5) / (wacc-g) + TV
        # For simplicity, use FCF1 / (WACC - avg_growth/2) as indicative
        avg_g = f"({rg_cell}/2+$B$37/2)"
        formula = f"=({fcf1}/($B$57-{avg_g})+B70)/B72"
        set_cell(ws, r, 2+j, formula, BLACK_FONT, fmt=PRICE_FMT, align=CENTER)

# Table 3: Beta vs Risk-Free Rate
R = 94
sub_header(ws, R, "表3: 每股价值 vs Beta & 无风险利率 / Price vs Beta & Risk-Free Rate", 1, 8)
R = 95
beta_vals = [0.8, 0.9, 1.0, 1.14, 1.3]
rf_vals = [0.015, 0.020, 0.022, 0.025, 0.030]

set_cell(ws, R, 1, "Beta \\ Rf", SUBHEADER_FONT, LIGHT_BLUE_FILL, align=CENTER)
for j, rf in enumerate(rf_vals):
    set_cell(ws, R, 2+j, rf, SUBHEADER_FONT, LIGHT_BLUE_FILL, PCT_FMT, CENTER)

for i, beta in enumerate(beta_vals):
    r = R + 1 + i
    set_cell(ws, r, 1, beta, BLUE_FONT, fmt=RATIO_FMT, align=CENTER)
    for j, rf_v in enumerate(rf_vals):
        rf_cell = f"${get_column_letter(2+j)}${R}"
        beta_cell = f"$A{r}"
        # Ke = Rf + Beta * ERP
        # WACC_new = Ke * eq_wt + Kd_at * debt_wt
        ke = f"({rf_cell}+{beta_cell}*{ERP})"
        wacc_new = f"({ke}*{EQ_WEIGHT:.4f}+{AFTER_TAX_DEBT:.4f}*{DEBT_WEIGHT:.4f})"
        
        pv_parts = []
        for k in range(5):
            fcf_col = get_column_letter(6+k)
            period = 0.5 + k
            pv_parts.append(f"{fcf_col}54/(1+{wacc_new})^{period}")
        sum_pv = "+".join(pv_parts)
        
        tv = f"(J54*(1+B37)/({wacc_new}-B37))/(1+{wacc_new})^4.5"
        formula = f"=({sum_pv}+{tv}+B70)/B72"
        set_cell(ws, r, 2+j, formula, BLACK_FONT, fmt=PRICE_FMT, align=CENTER)


# ============================================================
# WACC SHEET
# ============================================================
ws2 = wb.create_sheet("WACC")
ws2.sheet_properties.tabColor = "4472C4"

ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 18

R = 1
section_header(ws2, R, "加权平均资本成本计算 / WACC CALCULATION", 1, 3)

R = 2
sub_header(ws2, R, "权益成本 / COST OF EQUITY (CAPM)", 1, 3)
items_ke = [
    ("无风险利率 / Risk-Free Rate (10Y国债)", RF_RATE, PCT_FMT, "Source: 中国10年期国债收益率 2026-03"),
    ("Beta (5年月度 vs 沪深300)", BETA, RATIO_FMT, "Source: investing.com 2026-03"),
    ("股权风险溢价 / Equity Risk Premium", ERP, PCT_FMT, "Source: A股历史ERP均值 6.5%"),
    ("权益成本 / Cost of Equity", None, PCT_FMT, None),
]
for i, (label, val, fmt, cmt) in enumerate(items_ke):
    r = 3 + i
    set_cell(ws2, r, 1, label, SUBHEADER_FONT if val is None else BLACK_FONT, LIGHT_GRAY_FILL if val is None else None)
    if val is not None:
        set_cell(ws2, r, 2, val, BLUE_FONT, LIGHT_GREEN_FILL, fmt, CENTER, cmt)
    else:
        set_cell(ws2, r, 2, "=B3+B4*B5", RESULT_FONT, LIGHT_GRAY_FILL, fmt, CENTER)

R = 8
sub_header(ws2, R, "债务成本 / COST OF DEBT", 1, 3)
items_kd = [
    ("税前债务成本 / Pre-Tax Cost of Debt", PRE_TAX_DEBT, PCT_FMT, "Source: LPR + 信用利差 估计"),
    ("有效税率 / Tax Rate", TAX_RATE, PCT_FMT, "Source: 中国宏桥有效税率"),
    ("税后债务成本 / After-Tax Cost of Debt", None, PCT_FMT, None),
]
for i, (label, val, fmt, cmt) in enumerate(items_kd):
    r = 9 + i
    set_cell(ws2, r, 1, label, SUBHEADER_FONT if val is None else BLACK_FONT, LIGHT_GRAY_FILL if val is None else None)
    if val is not None:
        set_cell(ws2, r, 2, val, BLUE_FONT, LIGHT_GREEN_FILL, fmt, CENTER, cmt)
    else:
        set_cell(ws2, r, 2, "=B9*(1-B10)", RESULT_FONT, LIGHT_GRAY_FILL, fmt, CENTER)

R = 13
sub_header(ws2, R, "资本结构 / CAPITAL STRUCTURE", 1, 3)
set_cell(ws2, 14, 1, "权益权重 / Equity Weight", BLACK_FONT)
set_cell(ws2, 14, 2, f"=DCF!B11/(DCF!B11+DCF!B14)", GREEN_FONT, fmt=PCT_FMT, align=CENTER)
set_cell(ws2, 15, 1, "债务权重 / Debt Weight", BLACK_FONT)
set_cell(ws2, 15, 2, f"=DCF!B14/(DCF!B11+DCF!B14)", GREEN_FONT, fmt=PCT_FMT, align=CENTER)

R = 16
set_cell(ws2, R, 1, "WACC (加权平均资本成本)", Font(name="Microsoft YaHei", size=14, bold=True, color="FF0000"), YELLOW_FILL)
set_cell(ws2, R, 2, "=B6*B14+B11*B15", Font(name="Microsoft YaHei", size=14, bold=True, color="FF0000"), YELLOW_FILL, PCT_FMT, CENTER)

# ============================================================
# ALUMINUM PRICE SENSITIVITY SHEET
# ============================================================
ws3 = wb.create_sheet("AL_Price")
ws3.sheet_properties.tabColor = "C00000"

ws3.column_dimensions['A'].width = 38
for c in range(2, 10):
    ws3.column_dimensions[get_column_letter(c)].width = 18

# --- Title ---
R = 1
section_header(ws3, R, "SHFE Aluminum Price Sensitivity / SHFE", 1, 9)
R = 2
set_cell(ws3, R, 1, "Shanghai Futures Exchange Aluminum", BLACK_FONT)
ws3.merge_cells('A2:I2')

# --- Section 1: Production Volume Assumptions ---
R = 4
section_header(ws3, R, "/ PRODUCTION VOLUME (FY2024A)", 1, 9)

vol_data = [
    ("/ Al Alloy Sales (10k tons)", 583.7,
     "Source: 2024, 583.7"),
    ("/ Alumina Sales (10k tons)", 1092.1,
     "Source: 2024, 1092.1"),
    ("/ Al Processing Sales (10k tons)", 76.6,
     "Source: 2024, 76.6"),
    ("/ Total Al Equivalent Volume (10k tons)", None, None),
    ("", None, None),
    ("FY2024A / Avg Realized Al Price (RMB/ton)", 20500,
     "Source: ~20500/ton (2024) from revenue/volume"),
    ("FY2024A / Avg Realized Alumina Price (RMB/ton)", 3800,
     "Source: ~3800/ton (2024) estimated"),
    ("FY2024A (M) / Revenue", 156169,
     "Source: =DCF!E45"),
]

R = 5
for i, (label, val, cmt) in enumerate(vol_data):
    r = R + i
    set_cell(ws3, r, 1, label, SUBHEADER_FONT, LIGHT_GRAY_FILL)
    if val is not None:
        set_cell(ws3, r, 2, val, BLUE_FONT, LIGHT_GREEN_FILL,
                 NUM_FMT if isinstance(val, (int, float)) and val > 100 else NUM_FMT_1,
                 CENTER, cmt)
    elif "Total" in label:
        set_cell(ws3, r, 2, "=B5+B7*0.12", BLACK_FONT, fmt=NUM_FMT_1, align=CENTER,
                 comment_text="Al alloy + processing (already counted) + alumina as Al equivalent (~12%)")

# --- Section 2: Revenue by AL Price Scenario ---
R = 14
section_header(ws3, R, "/ REVENUE BY SHFE AL PRICE SCENARIO", 1, 9)

R = 15
sub_header(ws3, R, "/ Assumptions", 1, 9)

R = 16
set_cell(ws3, R, 1, " / Al revenue as % of total rev", SUBHEADER_FONT, LIGHT_GRAY_FILL)
set_cell(ws3, R, 2, 0.72, BLUE_FONT, LIGHT_GREEN_FILL, PCT_FMT, CENTER,
         "Source: Aluminum alloy + processing ~72% of total revenue, rest is alumina & other")
R = 17
set_cell(ws3, R, 1, " / Alumina rev follows Al price (correlation)", SUBHEADER_FONT, LIGHT_GRAY_FILL)
set_cell(ws3, R, 2, 0.60, BLUE_FONT, LIGHT_GREEN_FILL, PCT_FMT, CENTER,
         "Source: Alumina price historically ~60% correlated to aluminum price moves")
R = 18
set_cell(ws3, R, 1, "EBIT / EBIT sensitivity to Al price", SUBHEADER_FONT, LIGHT_GRAY_FILL)
set_cell(ws3, R, 2, 1.5, BLUE_FONT, LIGHT_GREEN_FILL, RATIO_FMT, CENTER,
         "Source: Operating leverage ~1.5x. 10% revenue change -> ~15% EBIT change")

R = 20
section_header(ws3, R, "SHFE / SHFE AL PRICE -> REVENUE & VALUATION IMPACT", 1, 9)

# Column headers: AL Prices
al_prices = [19000, 20000, 21000, 22000, 23000, 24000, 25000, 26000]
R = 21
set_cell(ws3, R, 1, "SHFE (RMB/ton)", SUBHEADER_FONT, LIGHT_BLUE_FILL, align=CENTER)
for j, p in enumerate(al_prices):
    set_cell(ws3, R, 2+j, p, SUBHEADER_FONT, LIGHT_BLUE_FILL, '#,##0', CENTER)

# Row 22: Al price change vs 2024 base
R = 22
set_cell(ws3, R, 1, " vs 2024 / Al Price Change vs Base", BLACK_FONT, LIGHT_GRAY_FILL)
for j, p in enumerate(al_prices):
    formula = f"={get_column_letter(2+j)}21/$B$10-1"
    # Actually reference the base price in B10 (row 10 has avg realized price 20500)
    formula = f"={get_column_letter(2+j)}21/B10-1"
    set_cell(ws3, R, 2+j, formula, BLACK_FONT, LIGHT_GRAY_FILL, PCT_FMT, CENTER)

# Row 23: Revenue impact (Al portion)
R = 23
set_cell(ws3, R, 1, " / Al Revenue Change", BLACK_FONT)
for j in range(len(al_prices)):
    col = get_column_letter(2+j)
    # Al rev change = price_change * al_rev_pct
    formula = f"={col}22*$B$16"
    set_cell(ws3, R, 2+j, formula, BLACK_FONT, fmt=PCT_FMT, align=CENTER)

# Row 24: Alumina revenue impact
R = 24
set_cell(ws3, R, 1, " / Alumina Revenue Change", BLACK_FONT)
for j in range(len(al_prices)):
    col = get_column_letter(2+j)
    # Alumina change = price_change * (1-al_pct) * correlation
    formula = f"={col}22*(1-$B$16)*$B$17"
    set_cell(ws3, R, 2+j, formula, BLACK_FONT, fmt=PCT_FMT, align=CENTER)

# Row 25: Total revenue change
R = 25
set_cell(ws3, R, 1, " / Total Revenue Change", SUBHEADER_FONT, YELLOW_FILL)
for j in range(len(al_prices)):
    col = get_column_letter(2+j)
    formula = f"={col}23+{col}24"
    set_cell(ws3, R, 2+j, formula, RESULT_FONT, YELLOW_FILL, PCT_FMT, CENTER)

# Row 26: Implied Revenue (¥M)
R = 26
set_cell(ws3, R, 1, "FY2025E (M) / Implied Revenue", SUBHEADER_FONT, LIGHT_GRAY_FILL)
for j in range(len(al_prices)):
    col = get_column_letter(2+j)
    # Base revenue * (1 + total change) * (1 + organic growth ex-price)
    formula = f"=$B$12*(1+{col}25)*(1+0.03)"
    set_cell(ws3, R, 2+j, formula, BLACK_FONT, LIGHT_GRAY_FILL, NUM_FMT, CENTER)

# Row 27: Implied EBIT
R = 27
set_cell(ws3, R, 1, "FY2025E EBIT (M) / Implied EBIT", SUBHEADER_FONT)
for j in range(len(al_prices)):
    col = get_column_letter(2+j)
    # EBIT = Revenue * base_ebit_margin * (1 + price_change * leverage)
    formula = f"={col}26*0.24*(1+{col}22*$B$18*0.5)"
    set_cell(ws3, R, 2+j, formula, BLACK_FONT, fmt=NUM_FMT, align=CENTER)

# Row 28: EBIT Margin
R = 28
set_cell(ws3, R, 1, "EBIT / Implied EBIT Margin", BLACK_FONT)
for j in range(len(al_prices)):
    col = get_column_letter(2+j)
    formula = f"={col}27/{col}26"
    set_cell(ws3, R, 2+j, formula, BLACK_FONT, fmt=PCT_FMT, align=CENTER)

# Row 29: FCF
R = 29
set_cell(ws3, R, 1, "FCF (M) / Implied Unlevered FCF", SUBHEADER_FONT)
for j in range(len(al_prices)):
    col = get_column_letter(2+j)
    # FCF = EBIT*(1-tax) + Rev*DA% - Rev*Capex%
    formula = f"={col}27*(1-DCF!$B$41)+{col}26*DCF!$B$38-{col}26*DCF!$B$39"
    set_cell(ws3, R, 2+j, formula, BLACK_FONT, fmt=NUM_FMT, align=CENTER)

# Row 31: Implied Share Price (via simplified perpetuity)
R = 31
section_header(ws3, R, " / IMPLIED SHARE PRICE BY AL PRICE", 1, 9)

R = 32
set_cell(ws3, R, 1, "SHFE (RMB/ton)", SUBHEADER_FONT, LIGHT_BLUE_FILL, align=CENTER)
for j, p in enumerate(al_prices):
    set_cell(ws3, R, 2+j, p, SUBHEADER_FONT, LIGHT_BLUE_FILL, '#,##0', CENTER)

# Implied price using FCF perpetuity with 5-year ramp
R = 33
set_cell(ws3, R, 1, " (Y1 FCF) / Implied Price (Y1 Perpetuity)", 
         Font(name="Microsoft YaHei", size=12, bold=True, color="FF0000"), YELLOW_FILL)
for j in range(len(al_prices)):
    col = get_column_letter(2+j)
    formula = f"=({col}29/(DCF!$B$57-DCF!$B$37)+DCF!B70)/DCF!B72"
    c = ws3.cell(row=R, column=2+j, value=formula)
    c.font = Font(name="Microsoft YaHei", size=12, bold=True, color="FF0000")
    c.fill = YELLOW_FILL
    c.number_format = PRICE_FMT
    c.alignment = CENTER
    c.border = THIN_BORDER

# Row 34: vs Current Price
R = 34
set_cell(ws3, R, 1, " / Upside vs Current Price", SUBHEADER_FONT, LIGHT_GRAY_FILL)
for j in range(len(al_prices)):
    col = get_column_letter(2+j)
    formula = f"={col}33/DCF!B9-1"
    set_cell(ws3, R, 2+j, formula, BLACK_FONT, LIGHT_GRAY_FILL, PCT_FMT, CENTER)

# --- Section 3: Dual-axis table (Al Price x EBIT Margin -> Share Price) ---
R = 37
section_header(ws3, R, " x EBIT -> / AL PRICE x EBIT MARGIN -> IMPLIED SHARE PRICE", 1, 9)

R = 38
ebit_margins_al = [0.18, 0.20, 0.22, 0.24, 0.26, 0.28]
al_prices_short = [20000, 22000, 23000, 24000, 25000, 26000]

set_cell(ws3, R, 1, "AL \\ EBIT%", SUBHEADER_FONT, LIGHT_BLUE_FILL, align=CENTER)
for j, em in enumerate(ebit_margins_al):
    set_cell(ws3, R, 2+j, em, SUBHEADER_FONT, LIGHT_BLUE_FILL, PCT_FMT, CENTER)

for i, al_p in enumerate(al_prices_short):
    r = R + 1 + i
    set_cell(ws3, r, 1, al_p, BLUE_FONT, fmt='#,##0', align=CENTER)
    for j, em_v in enumerate(ebit_margins_al):
        al_cell = f"$A{r}"
        em_cell = f"${get_column_letter(2+j)}${R}"
        # Revenue = Base_rev * (1 + (al_price/base_price - 1) * (al_pct + (1-al_pct)*corr)) * (1+organic)
        rev_factor = f"(1+({al_cell}/$B$10-1)*($B$16+(1-$B$16)*$B$17))*(1+0.03)"
        rev = f"($B$12*{rev_factor})"
        # FCF = rev*ebit*(1-tax) + rev*da% - rev*capex%
        fcf = f"({rev}*{em_cell}*(1-DCF!$B$41)+{rev}*DCF!$B$38-{rev}*DCF!$B$39)"
        # Price = (FCF/(WACC-TGR) - NetDebt) / Shares
        formula = f"=({fcf}/(DCF!$B$57-DCF!$B$37)+DCF!B70)/DCF!B72"
        set_cell(ws3, r, 2+j, formula, BLACK_FONT, fmt=PRICE_FMT, align=CENTER)

# Notes
R = 46
set_cell(ws3, R, 1, "/ Notes:", SUBHEADER_FONT)
R = 47
set_cell(ws3, R, 1, "1. 2024A ~20,500/ton", BLACK_FONT)
R = 48
set_cell(ws3, R, 1, "2. ~72%, ~28% / Al products ~72% of rev, alumina & other ~28%", BLACK_FONT)
R = 49
set_cell(ws3, R, 1, "3.  ~60% / Alumina price correlated ~60% to aluminum price moves", BLACK_FONT)
R = 50
set_cell(ws3, R, 1, "4. EBIT~1.5x / Operating leverage: 10% AL price up -> ~15% EBIT change", BLACK_FONT)
R = 51
set_cell(ws3, R, 1, "5. Y1 FCF / Perpetuity / Valuation uses Y1 FCF perpetuity method (simplified DCF)", BLACK_FONT)

# ============================================================
# SAVE
# ============================================================
output_path = r"c:\Users\rickylu\.gemini\antigravity\scratch\financial-services-plugins\002379_DCF_Model_2026-03-09.xlsx"
wb.save(output_path)
print(f"[OK] DCF Model saved to: {output_path}")
print(f"Company: {TICKER}")
print(f"Stock Price: {STOCK_PRICE}")
print(f"Market Cap: {MARKET_CAP:,.0f}M RMB")
print(f"WACC: {WACC:.2%}")
print(f"Shares: {SHARES_OUT:,}M")
