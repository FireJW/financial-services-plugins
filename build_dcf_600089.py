"""
特变电工 (600089.SH) DCF Valuation Model
输变电 + 新能源 + 煤炭 多元化集团
Generated: 2026-03-10
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
BF = Font(name="Microsoft YaHei", size=11, color="0000FF")
KF = Font(name="Microsoft YaHei", size=11, color="000000")
GF = Font(name="Microsoft YaHei", size=11, color="008000")
HF = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
SF = Font(name="Microsoft YaHei", size=11, bold=True, color="000000")
TF = Font(name="Microsoft YaHei", size=14, bold=True, color="000000")
RF = Font(name="Microsoft YaHei", size=12, bold=True, color="000000")
RB = Font(name="Microsoft YaHei", size=14, bold=True, color="FF0000")

DB = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
LB = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
LG = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
YL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

CA = Alignment(horizontal="center", vertical="center")
BD = Border(left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))
P='0.0%'; N='#,##0'; N1='#,##0.0'; PR='#,##0.00'

def sc(ws, r, c, v, f=None, fl=None, fm=None, cm=None):
    cell = ws.cell(row=r, column=c, value=v)
    if f: cell.font = f
    if fl: cell.fill = fl
    if fm: cell.number_format = fm
    cell.alignment = CA; cell.border = BD
    if cm: cell.comment = openpyxl.comments.Comment(cm, "DCF")
    return cell

def hdr(ws, r, t, c1=1, c2=10):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c); cell.fill = DB; cell.font = HF; cell.alignment = CA; cell.border = BD
    ws.cell(row=r, column=c1, value=t)
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

def shdr(ws, r, t, c1=1, c2=10):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c); cell.fill = LB; cell.font = SF; cell.alignment = CA; cell.border = BD
    ws.cell(row=r, column=c1, value=t)

# ============================================================
# DATA
# ============================================================
TICKER = "600089.SH"
DATE = "2026-03-10"

PRICE = 32.76
SHARES = 5053      # M shares
MCAP = PRICE * SHARES

hist = {
    "year":    [2021,    2022,     2023,     2024],
    "rev":     [72000,   95887,    98123,    97782],
    "ebit":    [19500,   24500,    16000,    10500],
    "ebit_pct":[0.271,   0.256,    0.163,    0.107],
    "tax":     [2900,    3700,     2400,     1600],
    "ni":      [13500,   15883,    10703,    4135],
    "da":      [4000,    4421,     5041,     12765],  # 2024 includes impairments
    "capex":   [6000,    8000,     9000,     9500],
    "ebitda":  [23500,   28921,    21041,    23265],
}

# Balance Sheet
DEBT = 55000       # M estimated interest-bearing debt
CASH = 21594       # M cash
NET_DEBT = DEBT - CASH  # 33,406 M

# WACC Inputs
Rf = 0.022
BETA = 0.63        # Low beta - diversified conglomerate
ERP = 0.065
KE = Rf + BETA * ERP
KD_PRE = 0.038
TAX_R = 0.15       # High-tech enterprise preferential rate
KD_POST = KD_PRE * (1 - TAX_R)
EV = MCAP + NET_DEBT
EQ_W = MCAP / EV
DT_W = NET_DEBT / EV
WACC_V = KE * EQ_W + KD_POST * DT_W

proj = {
    "years": [2025, 2026, 2027, 2028, 2029],
    "bear":  {"rg": [0.02, 0.02, 0.02, 0.01, 0.01], "em": [0.09, 0.09, 0.08, 0.08, 0.08]},
    "base":  {"rg": [0.05, 0.06, 0.05, 0.04, 0.03], "em": [0.12, 0.13, 0.13, 0.12, 0.12]},
    "bull":  {"rg": [0.08, 0.10, 0.08, 0.06, 0.05], "em": [0.16, 0.17, 0.17, 0.16, 0.15]},
}
DA_P = 0.055
CX_P = 0.095       # High capex due to polysilicon + coal mining
NW_P = 0.01
TG = {"bear": 0.02, "base": 0.025, "bull": 0.035}

# ============================================================
# DCF SHEET
# ============================================================
ws = wb.active; ws.title = "DCF"; ws.sheet_properties.tabColor = "17365D"
ws.column_dimensions['A'].width = 32
for c in range(2, 12): ws.column_dimensions[get_column_letter(c)].width = 16

R = 1
sc(ws, R, 1, f"TBEA (600089.SH) DCF Model", TF)
ws.merge_cells('A1:J1')
sc(ws, 2, 1, f"TICKER: {TICKER} | DATE: {DATE} | FY2024", KF)
ws.merge_cells('A2:J2')
sc(ws, 3, 1, "Unit: RMB Millions (M)", KF)
ws.merge_cells('A3:J3')

# Case Selector
hdr(ws, 5, "CASE SELECTOR", 1, 10)
sc(ws, 6, 1, "Active Case (1=Bear, 2=Base, 3=Bull)", SF, YL)
sc(ws, 6, 2, 2, BF, YL, cm="User input: 1/2/3")

# Market Data
hdr(ws, 8, "MARKET DATA", 1, 10)
mkt = [
    ("Stock Price", PRICE, PR, "Source: 2026-03-09"),
    ("Shares Outstanding (M)", SHARES, N, "Source: 50.53 yi"),
    ("Market Cap (M)", None, N, None),
    ("Total Debt (M)", DEBT, N, "Source: estimated from D/E"),
    ("Cash (M)", CASH, N, "Source: 2025Q3 report"),
    ("Net Debt (M)", None, N, None),
    ("Enterprise Value (M)", None, N, None),
]
for i, (lb, v, fm, cm) in enumerate(mkt):
    r = 9 + i
    sc(ws, r, 1, lb, SF, GY)
    if v is not None: sc(ws, r, 2, v, BF, LG, fm, cm)
    elif i == 2: sc(ws, r, 2, "=B9*B10", KF, fm=fm)
    elif i == 5: sc(ws, r, 2, "=B12-B13", KF, fm=fm)
    elif i == 6: sc(ws, r, 2, "=B11+B14", KF, fm=fm)

# Scenario Assumptions
hdr(ws, 17, "SCENARIO ASSUMPTIONS", 1, 10)
hp = ["Assumption"] + [f"FY{y}E" for y in proj["years"]]

def wc(ws, sr, name, ck, tgk):
    shdr(ws, sr, name, 1, 10)
    for c, h in enumerate(hp, 1): sc(ws, sr+1, c, h, SF, LB)
    r = sr + 2
    sc(ws, r, 1, "Revenue Growth", SF, GY)
    for j, v in enumerate(proj[ck]["rg"]): sc(ws, r, 2+j, v, BF, LG, P)
    r += 1
    sc(ws, r, 1, "EBIT Margin", SF, GY)
    for j, v in enumerate(proj[ck]["em"]): sc(ws, r, 2+j, v, BF, LG, P)
    r += 1
    sc(ws, r, 1, "Terminal Growth", SF, GY)
    sc(ws, r, 2, TG[tgk], BF, LG, P)

wc(ws, 18, "BEAR CASE", "bear", "bear")
wc(ws, 23, "BASE CASE", "base", "base")
wc(ws, 28, "BULL CASE", "bull", "bull")

# Selected Case
shdr(ws, 33, "SELECTED CASE (B6)", 1, 10)
for c, h in enumerate(hp, 1): sc(ws, 34, c, h, SF, LB)

sc(ws, 35, 1, "Revenue Growth", SF, YL)
for j in range(5):
    cl = get_column_letter(2+j)
    sc(ws, 35, 2+j, f'=IF($B$6=1,{cl}$20,IF($B$6=2,{cl}$25,{cl}$30))', KF, YL, P)

sc(ws, 36, 1, "EBIT Margin", SF, YL)
for j in range(5):
    cl = get_column_letter(2+j)
    sc(ws, 36, 2+j, f'=IF($B$6=1,{cl}$21,IF($B$6=2,{cl}$26,{cl}$31))', KF, YL, P)

sc(ws, 37, 1, "Terminal Growth", SF, YL)
sc(ws, 37, 2, '=IF($B$6=1,$B$22,IF($B$6=2,$B$27,$B$32))', KF, YL, P)

sc(ws, 38, 1, "D&A % Rev", SF, GY); sc(ws, 38, 2, DA_P, BF, LG, P, "~5.5% normalized (excl 2024 impairment)")
sc(ws, 39, 1, "CapEx % Rev", SF, GY); sc(ws, 39, 2, CX_P, BF, LG, P, "~9.5% (poly+coal heavy capex)")
sc(ws, 40, 1, "NWC % dRev", SF, GY); sc(ws, 40, 2, NW_P, BF, LG, P)
sc(ws, 41, 1, "Tax Rate", SF, GY); sc(ws, 41, 2, TAX_R, BF, LG, P, "15% HNTE preferential rate")

# P&L
hdr(ws, 43, "INCOME STATEMENT & FREE CASH FLOW", 1, 10)
yrs = [""] + [f"FY{y}A" for y in hist["year"]] + [f"FY{y}E" for y in proj["years"]]
for c, h in enumerate(yrs, 1):
    fl = LB if "A" in str(h) else GY if "E" in str(h) else None
    sc(ws, 44, c, h, SF, fl)

# Rev (45)
sc(ws, 45, 1, "Revenue", SF)
for i, rv in enumerate(hist["rev"]):
    sc(ws, 45, 2+i, rv, BF, fm=N, cm=f"FY{hist['year'][i]}")
for j in range(5):
    sc(ws, 45, 6+j, f"={get_column_letter(5+j)}45*(1+{get_column_letter(2+j)}$35)", KF, fm=N)

# Growth (46)
sc(ws, 46, 1, "  Growth %", KF)
for i in range(1, 4):
    sc(ws, 46, 2+i, f"={get_column_letter(2+i)}45/{get_column_letter(1+i)}45-1", KF, fm=P)
for j in range(5):
    sc(ws, 46, 6+j, f"={get_column_letter(6+j)}45/{get_column_letter(5+j)}45-1", KF, fm=P)

# EBIT (47)
sc(ws, 47, 1, "EBIT", SF)
for i, eb in enumerate(hist["ebit"]):
    sc(ws, 47, 2+i, eb, BF, fm=N)
for j in range(5):
    sc(ws, 47, 6+j, f"={get_column_letter(6+j)}45*{get_column_letter(2+j)}$36", KF, fm=N)

# EBIT% (48)
sc(ws, 48, 1, "  EBIT Margin", KF)
for i in range(4): sc(ws, 48, 2+i, f"={get_column_letter(2+i)}47/{get_column_letter(2+i)}45", KF, fm=P)
for j in range(5): sc(ws, 48, 6+j, f"={get_column_letter(6+j)}47/{get_column_letter(6+j)}45", KF, fm=P)

# Tax (49)
sc(ws, 49, 1, "(-) Taxes", KF)
for i, tx in enumerate(hist["tax"]): sc(ws, 49, 2+i, -abs(tx), BF, fm=N)
for j in range(5): sc(ws, 49, 6+j, f"=-{get_column_letter(6+j)}47*$B$41", KF, fm=N)

# NOPAT (50)
sc(ws, 50, 1, "NOPAT", SF)
for i in range(4): sc(ws, 50, 2+i, f"={get_column_letter(2+i)}47+{get_column_letter(2+i)}49", KF, fm=N)
for j in range(5): sc(ws, 50, 6+j, f"={get_column_letter(6+j)}47+{get_column_letter(6+j)}49", KF, fm=N)

# D&A (51)
sc(ws, 51, 1, "(+) D&A", KF)
for i, da in enumerate(hist["da"]): sc(ws, 51, 2+i, da, BF, fm=N)
for j in range(5): sc(ws, 51, 6+j, f"={get_column_letter(6+j)}45*$B$38", KF, fm=N)

# CapEx (52)
sc(ws, 52, 1, "(-) CapEx", KF)
for i, cx in enumerate(hist["capex"]): sc(ws, 52, 2+i, -abs(cx), BF, fm=N)
for j in range(5): sc(ws, 52, 6+j, f"=-{get_column_letter(6+j)}45*$B$39", KF, fm=N)

# NWC (53)
sc(ws, 53, 1, "(-) dNWC", KF)
for i in range(4): sc(ws, 53, 2+i, 0, BF, fm=N)
for j in range(5): sc(ws, 53, 6+j, f"=-({get_column_letter(6+j)}45-{get_column_letter(5+j)}45)*$B$40", KF, fm=N)

# FCF (54)
sc(ws, 54, 1, "Unlevered FCF", SF, GY)
for i in range(4):
    cl = get_column_letter(2+i)
    sc(ws, 54, 2+i, f"={cl}50+{cl}51+{cl}52+{cl}53", RF, GY, N)
for j in range(5):
    cl = get_column_letter(6+j)
    sc(ws, 54, 6+j, f"={cl}50+{cl}51+{cl}52+{cl}53", RF, GY, N)

# DCF
hdr(ws, 56, "DCF VALUATION", 1, 10)
sc(ws, 57, 1, "WACC", SF, YL); sc(ws, 57, 2, "=WACC!B16", GF, YL, P)
sc(ws, 58, 1, "Discount Period", SF)
for j in range(5): sc(ws, 58, 6+j, 0.5+j, BF, fm=N1)
sc(ws, 59, 1, "Discount Factor", KF)
for j in range(5): sc(ws, 59, 6+j, f"=1/(1+$B$57)^{get_column_letter(6+j)}58", KF, fm='0.0000')
sc(ws, 60, 1, "PV of FCF", SF)
for j in range(5):
    cl = get_column_letter(6+j)
    sc(ws, 60, 6+j, f"={cl}54*{cl}59", KF, fm=N)

sc(ws, 62, 1, "Terminal FCF", KF); sc(ws, 62, 2, "=J54*(1+B37)", KF, fm=N)
sc(ws, 63, 1, "Terminal Value", KF); sc(ws, 63, 2, "=B62/(B57-B37)", KF, fm=N)
sc(ws, 64, 1, "PV of Terminal Value", SF); sc(ws, 64, 2, "=B63/(1+B57)^4.5", KF, fm=N)

hdr(ws, 66, "VALUATION SUMMARY", 1, 10)
sc(ws, 67, 1, "Sum PV of FCFs", SF, GY); sc(ws, 67, 2, "=SUM(F60:J60)", KF, GY, N)
sc(ws, 68, 1, "PV of Terminal Value", SF, GY); sc(ws, 68, 2, "=B64", KF, GY, N)
sc(ws, 69, 1, "Enterprise Value", SF, YL); sc(ws, 69, 2, "=B67+B68", RF, YL, N)
sc(ws, 70, 1, "(-) Net Debt", SF, GY); sc(ws, 70, 2, "=-B14", KF, GY, N)
sc(ws, 71, 1, "Equity Value", SF, YL); sc(ws, 71, 2, "=B69+B70", RF, YL, N)
sc(ws, 72, 1, "Shares (M)", KF); sc(ws, 72, 2, "=B10", GF, fm=N)
sc(ws, 73, 1, "IMPLIED PRICE", RB, YL); sc(ws, 73, 2, "=B71/B72", RB, YL, PR)
sc(ws, 74, 1, "Current Price", SF); sc(ws, 74, 2, "=B9", GF, fm=PR)
sc(ws, 75, 1, "Implied Upside/(Downside)", SF, YL); sc(ws, 75, 2, "=B73/B74-1", RF, YL, P)

# Sensitivity 1: WACC vs TGR
hdr(ws, 77, "SENSITIVITY ANALYSIS", 1, 10)
shdr(ws, 78, "Table 1: Price vs WACC & Terminal Growth", 1, 8)
tgv = [0.015, 0.020, 0.025, 0.030, 0.035]
wv = [0.05, 0.06, 0.07, 0.08, 0.09]
sc(ws, 79, 1, "WACC \\ TGR", SF, LB)
for j, t in enumerate(tgv): sc(ws, 79, 2+j, t, SF, LB, P)
for i, w in enumerate(wv):
    r = 80+i
    sc(ws, r, 1, w, BF, fm=P)
    for j, t in enumerate(tgv):
        pvs = "+".join([f"{get_column_letter(6+k)}54/(1+$A{r})^{0.5+k}" for k in range(5)])
        tc = f"${get_column_letter(2+j)}$79"; wc = f"$A{r}"
        tv = f"(J54*(1+{tc})/({wc}-{tc}))/(1+{wc})^4.5"
        sc(ws, r, 2+j, f"=({pvs}+{tv}+B70)/B72", KF, fm=PR)

# Sensitivity 2: Rev Growth vs EBIT Margin
shdr(ws, 86, "Table 2: Price vs Rev Growth & EBIT Margin", 1, 8)
rgv = [0.02, 0.04, 0.06, 0.08, 0.10]
emv = [0.08, 0.10, 0.12, 0.14, 0.16]
sc(ws, 87, 1, "RevGr \\ EBIT%", SF, LB)
for j, e in enumerate(emv): sc(ws, 87, 2+j, e, SF, LB, P)
for i, rg in enumerate(rgv):
    r = 88+i
    sc(ws, r, 1, rg, BF, fm=P)
    for j, e in enumerate(emv):
        ec = f"${get_column_letter(2+j)}$87"; rc = f"$A{r}"
        fcf = f"(E45*(1+{rc})*{ec}*(1-$B$41)+E45*(1+{rc})*$B$38-E45*(1+{rc})*$B$39-E45*{rc}*$B$40)"
        ag = f"({rc}/2+$B$37/2)"
        sc(ws, r, 2+j, f"=({fcf}/($B$57-{ag})+B70)/B72", KF, fm=PR)

# Sensitivity 3: Beta vs Rf
shdr(ws, 94, "Table 3: Price vs Beta & Risk-Free Rate", 1, 8)
bv = [0.40, 0.50, 0.63, 0.75, 0.90]
rv = [0.015, 0.020, 0.022, 0.025, 0.030]
sc(ws, 95, 1, "Beta \\ Rf", SF, LB)
for j, rf in enumerate(rv): sc(ws, 95, 2+j, rf, SF, LB, P)
for i, b in enumerate(bv):
    r = 96+i
    sc(ws, r, 1, b, BF, fm='0.00')
    for j, rf in enumerate(rv):
        rc = f"${get_column_letter(2+j)}$95"; bc = f"$A{r}"
        ke = f"({rc}+{bc}*{ERP})"
        wn = f"({ke}*{EQ_W:.4f}+{KD_POST:.4f}*{DT_W:.4f})"
        pvs = "+".join([f"{get_column_letter(6+k)}54/(1+{wn})^{0.5+k}" for k in range(5)])
        tv = f"(J54*(1+B37)/({wn}-B37))/(1+{wn})^4.5"
        sc(ws, r, 2+j, f"=({pvs}+{tv}+B70)/B72", KF, fm=PR)

# ============================================================
# WACC SHEET
# ============================================================
ws2 = wb.create_sheet("WACC"); ws2.sheet_properties.tabColor = "4472C4"
ws2.column_dimensions['A'].width = 40; ws2.column_dimensions['B'].width = 18

hdr(ws2, 1, "WACC CALCULATION", 1, 3)
shdr(ws2, 2, "COST OF EQUITY (CAPM)", 1, 3)
sc(ws2, 3, 1, "Risk-Free Rate (10Y)", KF); sc(ws2, 3, 2, Rf, BF, LG, P, "China 10Y bond")
sc(ws2, 4, 1, "Beta (5Y monthly vs CSI300)", KF); sc(ws2, 4, 2, BETA, BF, LG, '0.00', "investing.com, low due to diversification")
sc(ws2, 5, 1, "Equity Risk Premium", KF); sc(ws2, 5, 2, ERP, BF, LG, P)
sc(ws2, 6, 1, "Cost of Equity", SF, GY); sc(ws2, 6, 2, "=B3+B4*B5", RF, GY, P)
shdr(ws2, 8, "COST OF DEBT", 1, 3)
sc(ws2, 9, 1, "Pre-Tax Cost of Debt", KF); sc(ws2, 9, 2, KD_PRE, BF, LG, P)
sc(ws2, 10, 1, "Tax Rate", KF); sc(ws2, 10, 2, TAX_R, BF, LG, P, "15% HNTE")
sc(ws2, 11, 1, "After-Tax Cost of Debt", SF, GY); sc(ws2, 11, 2, "=B9*(1-B10)", RF, GY, P)
shdr(ws2, 13, "CAPITAL STRUCTURE", 1, 3)
sc(ws2, 14, 1, "Equity Weight", KF); sc(ws2, 14, 2, "=DCF!B11/(DCF!B11+DCF!B14)", GF, fm=P)
sc(ws2, 15, 1, "Debt Weight", KF); sc(ws2, 15, 2, "=DCF!B14/(DCF!B11+DCF!B14)", GF, fm=P)
sc(ws2, 16, 1, "WACC", RB, YL); sc(ws2, 16, 2, "=B6*B14+B11*B15", RB, YL, P)

# ============================================================
# SEGMENT SHEET (Business Breakdown)
# ============================================================
ws3 = wb.create_sheet("Segments"); ws3.sheet_properties.tabColor = "548235"
ws3.column_dimensions['A'].width = 30
for c in range(2, 8): ws3.column_dimensions[get_column_letter(c)].width = 16

hdr(ws3, 1, "BUSINESS SEGMENT ANALYSIS (FY2024)", 1, 7)

# Revenue breakdown
shdr(ws3, 2, "Revenue by Segment (M RMB)", 1, 7)
sc(ws3, 3, 1, "Segment", SF, LB); sc(ws3, 3, 2, "Revenue", SF, LB)
sc(ws3, 3, 3, "% of Total", SF, LB); sc(ws3, 3, 4, "Key Driver", SF, LB)

segs = [
    ("Transformers / T&D", 42990, 0.4396, "Grid investment + exports"),
    ("Coal", 19260, 0.1970, "Coal price + volume"),
    ("New Energy + Poly", 18530, 0.1895, "Poly price (crashed in 2024)"),
    ("Power Generation", 5600, 0.0573, "Installed capacity"),
    ("New Materials", 5610, 0.0574, "Specialty products"),
    ("Other", 5792, 0.0592, ""),
]
for i, (nm, rv, pct, dr) in enumerate(segs):
    r = 4 + i
    sc(ws3, r, 1, nm, SF, GY); sc(ws3, r, 2, rv, BF, fm=N, cm=f"Source: FY2024 annual")
    sc(ws3, r, 3, pct, KF, fm=P); sc(ws3, r, 4, dr, KF)

sc(ws3, 10, 1, "Total Revenue", SF, YL); sc(ws3, 10, 2, "=SUM(B4:B9)", RF, YL, N)
sc(ws3, 10, 3, "=SUM(C4:C9)", RF, YL, P)

# Key notes
sc(ws3, 12, 1, "KEY NOTES:", SF)
sc(ws3, 13, 1, "1. Transformer business is the core growth driver (grid investment)", KF)
sc(ws3, 14, 1, "2. Poly silicon prices crashed ~70% in 2024, dragging total profitability", KF)
sc(ws3, 15, 1, "3. Coal segment stable but mature, ~20% of rev", KF)
sc(ws3, 16, 1, "4. NI fell from 159B (2022) to 41B (2024) mainly due to polysilicon", KF)
sc(ws3, 17, 1, "5. EBIT margin drop: 25.6% (2022) -> 10.7% (2024)", KF)
sc(ws3, 18, 1, "6. Recovery thesis: poly price stabilization + transformer order backlog", KF)

# ============================================================
# SAVE
# ============================================================
out = r"c:\Users\rickylu\.gemini\antigravity\scratch\financial-services-plugins\600089_DCF_Model_2026-03-10.xlsx"
wb.save(out)
print(f"[OK] Saved: {out}")
print(f"Ticker: {TICKER}")
print(f"Price: {PRICE}")
print(f"MCap: {MCAP:,.0f}M")
print(f"WACC: {WACC_V:.2%}")
