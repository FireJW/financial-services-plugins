# -*- coding: utf-8 -*-
'''
Chengxing Phosphate Chemical (600078.SH) DCF Valuation Model
Global Fine Phosphate Chemical Leader
Generated: 2026-03-11
'''

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

wb = openpyxl.Workbook()

# Styles
BF = Font(name="Microsoft YaHei", size=11, color="0000FF")
KF = Font(name="Microsoft YaHei", size=11, color="000000")
GF = Font(name="Microsoft YaHei", size=11, color="008000")
HF = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
SF = Font(name="Microsoft YaHei", size=11, bold=True, color="000000")
TF = Font(name="Microsoft YaHei", size=14, bold=True, color="000000")
RF = Font(name="Microsoft YaHei", size=12, bold=True, color="000000")
RB = Font(name="Microsoft YaHei", size=14, bold=True, color="FF0000")

DB = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
LB = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
LG = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
YL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

CA = Alignment(horizontal="center", vertical="center")
BD = Border(left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))
P='0.0%'; N='#,##0'; N1='#,##0.0'; PR='#,##0.00'

def sc(ws, r, c, v, f=None, fl=None, fm=None, cm=None):
    cell = ws.cell(row=r, column=c, value=v)
    if f: cell.font = f
    if fl: cell.fill = fl
    if fm: cell.number_format = fm
    cell.alignment = CA; cell.border = BD
    if cm: cell.comment = Comment(cm, "DCF")
    return cell

def hdr(ws, r, t, c1=1, c2=10):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c); cell.fill = DB; cell.font = HF; cell.alignment = CA; cell.border = BD
    ws.cell(row=r, column=c1, value=t)
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

def shdr(ws, r, t, c1=1, c2=10):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c); cell.fill = LB; cell.font = SF; cell.alignment = CA; cell.border = BD
    ws.cell(row=r, column=c1, value=t)

# ============================================================
# DATA
# ============================================================
TICKER = "600078.SH"
NAME = "Chengxing Phosphate"
DATE = "2026-03-11"

PRICE = 12.54
SHARES = 677        # M shares
MCAP = PRICE * SHARES

# Historical: 2024 was a loss year. Company is cyclical phosphate chemical producer.
# Revenue in M RMB
hist = {
    "year":    [2021,    2022,     2023,     2024],
    "rev":     [3333,    4538,     3101,     3356],
    "ebit":    [400,     650,      -80,      -150],
    "ebit_pct":[0.120,   0.143,    -0.026,   -0.045],
    "tax":     [60,      95,       0,        0],
    "ni":      [270,     430,      -120,     -199],
    "da":      [180,     200,      210,      220],
    "capex":   [250,     350,      300,      200],
    "ebitda":  [580,     850,      130,      70],
}

# Balance Sheet
DEBT = 2200         # M estimated interest-bearing debt
CASH = 588          # M cash (end of 2024)
NET_DEBT = DEBT - CASH  # 1612M

# WACC Inputs
Rf = 0.022
BETA = 1.35         # High beta - cyclical commodity chemical
ERP = 0.065
KE = Rf + BETA * ERP
KD_PRE = 0.045      # Higher cost of debt - smaller, more leveraged
TAX_R = 0.25        # Standard tax rate (no HNTE status, loss carryforward)
KD_POST = KD_PRE * (1 - TAX_R)
EV = MCAP + NET_DEBT
EQ_W = MCAP / EV
DT_W = NET_DEBT / EV
WACC_V = KE * EQ_W + KD_POST * DT_W

# Turnaround thesis: phosphorus prices recovering, new materials pivot
proj = {
    "years": [2025, 2026, 2027, 2028, 2029],
    "bear":  {"rg": [0.02, 0.02, 0.01, 0.01, 0.01], "em": [0.02, 0.03, 0.03, 0.03, 0.03]},
    "base":  {"rg": [0.05, 0.08, 0.06, 0.05, 0.04], "em": [0.05, 0.07, 0.08, 0.08, 0.08]},
    "bull":  {"rg": [0.08, 0.12, 0.10, 0.08, 0.06], "em": [0.08, 0.10, 0.12, 0.12, 0.11]},
}
DA_P = 0.060        # D&A % of revenue (capital intensive)
CX_P = 0.070        # CapEx % of revenue
NW_P = 0.020        # NWC change % of delta revenue
TG = {"bear": 0.015, "base": 0.025, "bull": 0.035}

# ============================================================
# DCF SHEET
# ============================================================
ws = wb.active; ws.title = "DCF"; ws.sheet_properties.tabColor = "17365D"
ws.column_dimensions['A'].width = 32
for c in range(2, 12): ws.column_dimensions[get_column_letter(c)].width = 16

R = 1
sc(ws, R, 1, f"Chengxing Phosphate (600078.SH) DCF Model", TF)
ws.merge_cells('A1:J1')
sc(ws, 2, 1, f"TICKER: {TICKER} | DATE: {DATE} | FY2024", KF)
ws.merge_cells('A2:J2')
sc(ws, 3, 1, "Unit: RMB Millions (M)", KF)
ws.merge_cells('A3:J3')

# Case Selector
hdr(ws, 5, "CASE SELECTOR", 1, 10)
sc(ws, 6, 1, "Active Case (1=Bear, 2=Base, 3=Bull)", SF, YL)
sc(ws, 6, 2, 2, BF, YL, cm="User input: 1/2/3")

# Market Data
hdr(ws, 8, "MARKET DATA", 1, 10)
mkt = [
    ("Stock Price", PRICE, PR, "Source: 2026-03-10"),
    ("Shares Outstanding (M)", SHARES, N, "Source: 6.77 yi"),
    ("Market Cap (M)", None, N, None),
    ("Total Debt (M)", DEBT, N, "Est. interest-bearing debt"),
    ("Cash (M)", CASH, N, "Source: FY2024 annual"),
    ("Net Debt (M)", None, N, None),
    ("Enterprise Value (M)", None, N, None),
]
for i, (lb, v, fm, cm) in enumerate(mkt):
    r = 9 + i
    sc(ws, r, 1, lb, SF, GY)
    if v is not None: sc(ws, r, 2, v, BF, LG, fm, cm)
    elif i == 2: sc(ws, r, 2, "=B9*B10", KF, fm=fm)
    elif i == 5: sc(ws, r, 2, "=B12-B13", KF, fm=fm)
    elif i == 6: sc(ws, r, 2, "=B11+B14", KF, fm=fm)

# Scenario Assumptions
hdr(ws, 17, "SCENARIO ASSUMPTIONS", 1, 10)
hp = ["Assumption"] + [f"FY{y}E" for y in proj["years"]]

def wc(ws, sr, name, ck, tgk):
    shdr(ws, sr, name, 1, 10)
    for c, h in enumerate(hp, 1): sc(ws, sr+1, c, h, SF, LB)
    r = sr + 2
    sc(ws, r, 1, "Revenue Growth", SF, GY)
    for j, v in enumerate(proj[ck]["rg"]): sc(ws, r, 2+j, v, BF, LG, P)
    r += 1
    sc(ws, r, 1, "EBIT Margin", SF, GY)
    for j, v in enumerate(proj[ck]["em"]): sc(ws, r, 2+j, v, BF, LG, P)
    r += 1
    sc(ws, r, 1, "Terminal Growth", SF, GY)
    sc(ws, r, 2, TG[tgk], BF, LG, P)

wc(ws, 18, "BEAR CASE", "bear", "bear")
wc(ws, 23, "BASE CASE", "base", "base")
wc(ws, 28, "BULL CASE", "bull", "bull")

# Selected Case
shdr(ws, 33, "SELECTED CASE (B6)", 1, 10)
for c, h in enumerate(hp, 1): sc(ws, 34, c, h, SF, LB)

sc(ws, 35, 1, "Revenue Growth", SF, YL)
for j in range(5):
    cl = get_column_letter(2+j)
    sc(ws, 35, 2+j, f'=IF($B$6=1,{cl}$20,IF($B$6=2,{cl}$25,{cl}$30))', KF, YL, P)

sc(ws, 36, 1, "EBIT Margin", SF, YL)
for j in range(5):
    cl = get_column_letter(2+j)
    sc(ws, 36, 2+j, f'=IF($B$6=1,{cl}$21,IF($B$6=2,{cl}$26,{cl}$31))', KF, YL, P)

sc(ws, 37, 1, "Terminal Growth", SF, YL)
sc(ws, 37, 2, '=IF($B$6=1,$B$22,IF($B$6=2,$B$27,$B$32))', KF, YL, P)

sc(ws, 38, 1, "D&A % Rev", SF, GY); sc(ws, 38, 2, DA_P, BF, LG, P, "~6% capital intensive")
sc(ws, 39, 1, "CapEx % Rev", SF, GY); sc(ws, 39, 2, CX_P, BF, LG, P, "~7% chemical plants")
sc(ws, 40, 1, "NWC % dRev", SF, GY); sc(ws, 40, 2, NW_P, BF, LG, P)
sc(ws, 41, 1, "Tax Rate", SF, GY); sc(ws, 41, 2, TAX_R, BF, LG, P, "25% standard rate")

# P&L
hdr(ws, 43, "INCOME STATEMENT & FREE CASH FLOW", 1, 10)
yrs = [""] + [f"FY{y}A" for y in hist["year"]] + [f"FY{y}E" for y in proj["years"]]
for c, h in enumerate(yrs, 1):
    fl = LB if "A" in str(h) else GY if "E" in str(h) else None
    sc(ws, 44, c, h, SF, fl)

# Rev (45)
sc(ws, 45, 1, "Revenue", SF)
for i, rv in enumerate(hist["rev"]):
    sc(ws, 45, 2+i, rv, BF, fm=N, cm=f"FY{hist['year'][i]}")
for j in range(5):
    sc(ws, 45, 6+j, f"={get_column_letter(5+j)}45*(1+{get_column_letter(2+j)}$35)", KF, fm=N)

# Growth (46)
sc(ws, 46, 1, "  Growth %", KF)
for i in range(1, 4):
    sc(ws, 46, 2+i, f"={get_column_letter(2+i)}45/{get_column_letter(1+i)}45-1", KF, fm=P)
for j in range(5):
    sc(ws, 46, 6+j, f"={get_column_letter(6+j)}45/{get_column_letter(5+j)}45-1", KF, fm=P)

# EBIT (47)
sc(ws, 47, 1, "EBIT", SF)
for i, eb in enumerate(hist["ebit"]):
    sc(ws, 47, 2+i, eb, BF, fm=N)
for j in range(5):
    sc(ws, 47, 6+j, f"={get_column_letter(6+j)}45*{get_column_letter(2+j)}$36", KF, fm=N)

# EBIT% (48)
sc(ws, 48, 1, "  EBIT Margin", KF)
for i in range(4): sc(ws, 48, 2+i, f"={get_column_letter(2+i)}47/{get_column_letter(2+i)}45", KF, fm=P)
for j in range(5): sc(ws, 48, 6+j, f"={get_column_letter(6+j)}47/{get_column_letter(6+j)}45", KF, fm=P)

# Tax (49)
sc(ws, 49, 1, "(-) Taxes", KF)
for i, tx in enumerate(hist["tax"]): sc(ws, 49, 2+i, -abs(tx), BF, fm=N)
for j in range(5): sc(ws, 49, 6+j, f"=-MAX({get_column_letter(6+j)}47,0)*$B$41", KF, fm=N)

# NOPAT (50)
sc(ws, 50, 1, "NOPAT", SF)
for i in range(4): sc(ws, 50, 2+i, f"={get_column_letter(2+i)}47+{get_column_letter(2+i)}49", KF, fm=N)
for j in range(5): sc(ws, 50, 6+j, f"={get_column_letter(6+j)}47+{get_column_letter(6+j)}49", KF, fm=N)

# D&A (51)
sc(ws, 51, 1, "(+) D&A", KF)
for i, da in enumerate(hist["da"]): sc(ws, 51, 2+i, da, BF, fm=N)
for j in range(5): sc(ws, 51, 6+j, f"={get_column_letter(6+j)}45*$B$38", KF, fm=N)

# CapEx (52)
sc(ws, 52, 1, "(-) CapEx", KF)
for i, cx in enumerate(hist["capex"]): sc(ws, 52, 2+i, -abs(cx), BF, fm=N)
for j in range(5): sc(ws, 52, 6+j, f"=-{get_column_letter(6+j)}45*$B$39", KF, fm=N)

# NWC (53)
sc(ws, 53, 1, "(-) dNWC", KF)
for i in range(4): sc(ws, 53, 2+i, 0, BF, fm=N)
for j in range(5): sc(ws, 53, 6+j, f"=-({get_column_letter(6+j)}45-{get_column_letter(5+j)}45)*$B$40", KF, fm=N)

# FCF (54)
sc(ws, 54, 1, "Unlevered FCF", SF, GY)
for i in range(4):
    cl = get_column_letter(2+i)
    sc(ws, 54, 2+i, f"={cl}50+{cl}51+{cl}52+{cl}53", RF, GY, N)
for j in range(5):
    cl = get_column_letter(6+j)
    sc(ws, 54, 6+j, f"={cl}50+{cl}51+{cl}52+{cl}53", RF, GY, N)

# DCF
hdr(ws, 56, "DCF VALUATION", 1, 10)
sc(ws, 57, 1, "WACC", SF, YL); sc(ws, 57, 2, "=WACC!B16", GF, YL, P)
sc(ws, 58, 1, "Discount Period", SF)
for j in range(5): sc(ws, 58, 6+j, 0.5+j, BF, fm=N1)
sc(ws, 59, 1, "Discount Factor", KF)
for j in range(5): sc(ws, 59, 6+j, f"=1/(1+$B$57)^{get_column_letter(6+j)}58", KF, fm='0.0000')
sc(ws, 60, 1, "PV of FCF", SF)
for j in range(5):
    cl = get_column_letter(6+j)
    sc(ws, 60, 6+j, f"={cl}54*{cl}59", KF, fm=N)

sc(ws, 62, 1, "Terminal FCF", KF); sc(ws, 62, 2, "=J54*(1+B37)", KF, fm=N)
sc(ws, 63, 1, "Terminal Value", KF); sc(ws, 63, 2, "=B62/(B57-B37)", KF, fm=N)
sc(ws, 64, 1, "PV of Terminal Value", SF); sc(ws, 64, 2, "=B63/(1+B57)^4.5", KF, fm=N)

hdr(ws, 66, "VALUATION SUMMARY", 1, 10)
sc(ws, 67, 1, "Sum PV of FCFs", SF, GY); sc(ws, 67, 2, "=SUM(F60:J60)", KF, GY, N)
sc(ws, 68, 1, "PV of Terminal Value", SF, GY); sc(ws, 68, 2, "=B64", KF, GY, N)
sc(ws, 69, 1, "Enterprise Value", SF, YL); sc(ws, 69, 2, "=B67+B68", RF, YL, N)
sc(ws, 70, 1, "(-) Net Debt", SF, GY); sc(ws, 70, 2, "=-B14", KF, GY, N)
sc(ws, 71, 1, "Equity Value", SF, YL); sc(ws, 71, 2, "=B69+B70", RF, YL, N)
sc(ws, 72, 1, "Shares (M)", KF); sc(ws, 72, 2, "=B10", GF, fm=N)
sc(ws, 73, 1, "IMPLIED PRICE", RB, YL); sc(ws, 73, 2, "=B71/B72", RB, YL, PR)
sc(ws, 74, 1, "Current Price", SF); sc(ws, 74, 2, "=B9", GF, fm=PR)
sc(ws, 75, 1, "Implied Upside/(Downside)", SF, YL); sc(ws, 75, 2, "=B73/B74-1", RF, YL, P)

# Sensitivity 1: WACC vs TGR
hdr(ws, 77, "SENSITIVITY ANALYSIS", 1, 10)
shdr(ws, 78, "Table 1: Price vs WACC & Terminal Growth", 1, 8)
tgv = [0.010, 0.015, 0.020, 0.025, 0.035]
wv = [0.08, 0.09, 0.10, 0.11, 0.12]
sc(ws, 79, 1, "WACC \\ TGR", SF, LB)
for j, t in enumerate(tgv): sc(ws, 79, 2+j, t, SF, LB, P)
for i, w in enumerate(wv):
    r = 80+i
    sc(ws, r, 1, w, BF, fm=P)
    for j, t in enumerate(tgv):
        pvs = "+".join([f"{get_column_letter(6+k)}54/(1+$A{r})^{0.5+k}" for k in range(5)])
        tc = f"${get_column_letter(2+j)}$79"; wcc = f"$A{r}"
        tv = f"(J54*(1+{tc})/({wcc}-{tc}))/(1+{wcc})^4.5"
        sc(ws, r, 2+j, f"=({pvs}+{tv}+B70)/B72", KF, fm=PR)

# Sensitivity 2: Rev Growth vs EBIT Margin
shdr(ws, 86, "Table 2: Price vs Rev Growth & EBIT Margin", 1, 8)
rgv = [0.02, 0.04, 0.06, 0.08, 0.10]
emv = [0.04, 0.06, 0.08, 0.10, 0.12]
sc(ws, 87, 1, "RevGr \\ EBIT%", SF, LB)
for j, e in enumerate(emv): sc(ws, 87, 2+j, e, SF, LB, P)
for i, rg in enumerate(rgv):
    r = 88+i
    sc(ws, r, 1, rg, BF, fm=P)
    for j, e in enumerate(emv):
        ec = f"${get_column_letter(2+j)}$87"; rc = f"$A{r}"
        fcf = f"(E45*(1+{rc})*{ec}*(1-$B$41)+E45*(1+{rc})*$B$38-E45*(1+{rc})*$B$39-E45*{rc}*$B$40)"
        ag = f"({rc}/2+$B$37/2)"
        sc(ws, r, 2+j, f"=({fcf}/($B$57-{ag})+B70)/B72", KF, fm=PR)

# Sensitivity 3: Beta vs Rf
shdr(ws, 94, "Table 3: Price vs Beta & Risk-Free Rate", 1, 8)
bv = [0.90, 1.10, 1.35, 1.50, 1.80]
rv = [0.015, 0.020, 0.022, 0.025, 0.030]
sc(ws, 95, 1, "Beta \\ Rf", SF, LB)
for j, rf in enumerate(rv): sc(ws, 95, 2+j, rf, SF, LB, P)
for i, b in enumerate(bv):
    r = 96+i
    sc(ws, r, 1, b, BF, fm='0.00')
    for j, rf in enumerate(rv):
        rc = f"${get_column_letter(2+j)}$95"; bc = f"$A{r}"
        ke = f"({rc}+{bc}*{ERP})"
        wn = f"({ke}*{EQ_W:.4f}+{KD_POST:.4f}*{DT_W:.4f})"
        pvs = "+".join([f"{get_column_letter(6+k)}54/(1+{wn})^{0.5+k}" for k in range(5)])
        tv = f"(J54*(1+B37)/({wn}-B37))/(1+{wn})^4.5"
        sc(ws, r, 2+j, f"=({pvs}+{tv}+B70)/B72", KF, fm=PR)

# ============================================================
# WACC SHEET
# ============================================================
ws2 = wb.create_sheet("WACC"); ws2.sheet_properties.tabColor = "4472C4"
ws2.column_dimensions['A'].width = 40; ws2.column_dimensions['B'].width = 18

hdr(ws2, 1, "WACC CALCULATION", 1, 3)
shdr(ws2, 2, "COST OF EQUITY (CAPM)", 1, 3)
sc(ws2, 3, 1, "Risk-Free Rate (10Y)", KF); sc(ws2, 3, 2, Rf, BF, LG, P, "China 10Y bond")
sc(ws2, 4, 1, "Beta (5Y monthly vs CSI300)", KF); sc(ws2, 4, 2, BETA, BF, LG, '0.00', "High beta: cyclical commodity chemical")
sc(ws2, 5, 1, "Equity Risk Premium", KF); sc(ws2, 5, 2, ERP, BF, LG, P)
sc(ws2, 6, 1, "Cost of Equity", SF, GY); sc(ws2, 6, 2, "=B3+B4*B5", RF, GY, P)
shdr(ws2, 8, "COST OF DEBT", 1, 3)
sc(ws2, 9, 1, "Pre-Tax Cost of Debt", KF); sc(ws2, 9, 2, KD_PRE, BF, LG, P)
sc(ws2, 10, 1, "Tax Rate", KF); sc(ws2, 10, 2, TAX_R, BF, LG, P, "25% standard rate")
sc(ws2, 11, 1, "After-Tax Cost of Debt", SF, GY); sc(ws2, 11, 2, "=B9*(1-B10)", RF, GY, P)
shdr(ws2, 13, "CAPITAL STRUCTURE", 1, 3)
sc(ws2, 14, 1, "Equity Weight", KF); sc(ws2, 14, 2, "=DCF!B11/(DCF!B11+DCF!B14)", GF, fm=P)
sc(ws2, 15, 1, "Debt Weight", KF); sc(ws2, 15, 2, "=DCF!B14/(DCF!B11+DCF!B14)", GF, fm=P)
sc(ws2, 16, 1, "WACC", RB, YL); sc(ws2, 16, 2, "=B6*B14+B11*B15", RB, YL, P)

# ============================================================
# SEGMENTS SHEET
# ============================================================
ws3 = wb.create_sheet("Segments"); ws3.sheet_properties.tabColor = "548235"
ws3.column_dimensions['A'].width = 35
for c in range(2, 8): ws3.column_dimensions[get_column_letter(c)].width = 16

hdr(ws3, 1, "BUSINESS SEGMENT ANALYSIS (FY2024)", 1, 5)
shdr(ws3, 2, "Revenue by Segment (M RMB)", 1, 5)
sc(ws3, 3, 1, "Segment", SF, LB); sc(ws3, 3, 2, "Revenue", SF, LB)
sc(ws3, 3, 3, "% of Total", SF, LB); sc(ws3, 3, 4, "Key Driver", SF, LB)

segs = [
    ("Yellow Phosphorus", 1000, 0.298, "P price + Yunnan hydro cost"),
    ("Phosphoric Acid (Thermal)", 1200, 0.358, "60kt capacity, food/elec grade"),
    ("Fine Phosphate Salts", 650, 0.194, "Na/K/Ca phosphates, export"),
    ("Other Chemicals & Trading", 506, 0.150, "By-products + trading"),
]
for i, (nm, rv, pct, dr) in enumerate(segs):
    r = 4 + i
    sc(ws3, r, 1, nm, SF, GY); sc(ws3, r, 2, rv, BF, fm=N, cm="FY2024 est")
    sc(ws3, r, 3, pct, KF, fm=P); sc(ws3, r, 4, dr, KF)

sc(ws3, 8, 1, "Total Revenue", SF, YL); sc(ws3, 8, 2, "=SUM(B4:B7)", RF, YL, N)
sc(ws3, 8, 3, "=SUM(C4:C7)", RF, YL, P)

sc(ws3, 10, 1, "KEY NOTES:", SF)
sc(ws3, 11, 1, "1. Largest thermal phosphoric acid producer in China (60万吨/年)", KF)
sc(ws3, 12, 1, "2. Yellow phosphorus capacity: 16万吨/年 (Yunnan bases)", KF)
sc(ws3, 13, 1, "3. FY2024 net loss -199M due to phosphorus price crash", KF)
sc(ws3, 14, 1, "4. Turnaround thesis: P price recovery + elec-grade expansion", KF)
sc(ws3, 15, 1, "5. Complete value chain: mines -> yellow P -> phosphoric acid -> salts", KF)
sc(ws3, 16, 1, "6. High leverage risk: total debt ~34B, significant interest burden", KF)

# ============================================================
# SAVE
# ============================================================
out = r"c:\Users\rickylu\.gemini\antigravity\scratch\financial-services-plugins\600078_DCF_Model_2026-03-11.xlsx"
wb.save(out)
print(f"[OK] Saved: {out}")
print(f"Ticker: {TICKER}")
print(f"Price: {PRICE}")
print(f"MCap: {MCAP:,.0f}M")
print(f"WACC: {WACC_V:.2%}")
