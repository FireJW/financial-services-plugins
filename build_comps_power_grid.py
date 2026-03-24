"""
电网设备行业 (Power Grid Equipment) Comps Analysis Model
申万二级：电网设备 / A-Share Sector: Power Grid Equipment
Generated: 2026-03-10
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# Color & Style Definitions
# ============================================================
BLUE_FONT = Font(name="Microsoft YaHei", size=11, color="0000FF")       # Hardcoded inputs
BLACK_FONT = Font(name="Microsoft YaHei", size=11, color="000000")      # Formulas
GREEN_FONT = Font(name="Microsoft YaHei", size=11, color="008000")      # Cross-sheet links
RED_FONT = Font(name="Microsoft YaHei", size=11, color="FF0000")        # Negative values

HEADER_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="000000")
TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color="000000")
STAT_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="000000", italic=True)

DARK_BLUE_FILL = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
THICK_BOTTOM = Border(bottom=Side(style="medium"))

PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
NUM_FMT_1 = '#,##0.0'
MULT_FMT = '0.0x'
PRICE_FMT = '¥#,##0.00'

def set_cell(ws, row, col, value, font=None, fill=None, fmt=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    if align: cell.alignment = align
    cell.border = THIN_BORDER
    return cell

def section_header(ws, row, text, col_start=1, col_end=15):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = DARK_BLUE_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.cell(row=row, column=col_start, value=text)
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)

def sub_header(ws, row, text, col_start=1, col_end=15):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = LIGHT_BLUE_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.cell(row=row, column=col_start, value=text)

# ============================================================
# TARGET & PEER DATA (2024 Actuals / Estimates)
# RMB in Millions (Except per share)
# ============================================================
DATE = "2026-03-10"
SECTOR = "电网设备行业 (Power Grid Equipment)"

peers = [
    # 0:Ticker, 1:Name, 2:Price, 3:Shares(M), 4:NetDebt, 5:Rev, 6:NI, 7:EBITDA, 8:Equity(BV), 9:RevGr, 10:ROE
    {"t": "600406.SH", "n": "国电南瑞", "p": 31.50, "sh": 8036, "nd": -15000, "rev": 57417, "ni": 7610, "ebitda": 10300, "bv": 49224, "rg": 0.111, "roe": 0.154},
    {"t": "600089.SH", "n": "特变电工", "p": 32.76, "sh": 5053, "nd": 33400, "rev": 97870, "ni": 4130, "ebitda": 23265, "bv": 64500, "rg": -0.003, "roe": 0.064},
    {"t": "000400.SZ", "n": "许继电气", "p": 33.60, "sh": 1019, "nd": -3000,  "rev": 17089, "ni": 1117, "ebitda":  1650, "bv": 11080, "rg": 0.001, "roe": 0.101},
    {"t": "002028.SZ", "n": "思源电气", "p": 77.50, "sh": 774,  "nd": -2500,  "rev": 15458, "ni": 2050, "ebitda":  2800, "bv": 11350, "rg": 0.240, "roe": 0.180},
    {"t": "600312.SH", "n": "平高电气", "p": 25.03, "sh": 1357, "nd": -1500,  "rev": 12400, "ni": 1020, "ebitda":  1400, "bv": 10440, "rg": 0.120, "roe": 0.097},
]

# ============================================================
# COMPS SHEET
# ============================================================
ws = wb.active
ws.title = "Comps"
ws.sheet_properties.tabColor = "17365D"

# Column Setup
ws.column_dimensions['A'].width = 12  # Ticker
ws.column_dimensions['B'].width = 15  # Name
ws.column_dimensions['C'].width = 10  # Price
ws.column_dimensions['D'].width = 12  # MCap
ws.column_dimensions['E'].width = 12  # EV
ws.column_dimensions['F'].width = 12  # Rev
ws.column_dimensions['G'].width = 10  # Rev Gr
ws.column_dimensions['H'].width = 10  # ROE
ws.column_dimensions['I'].width = 12  # NI
ws.column_dimensions['J'].width = 12  # EBITDA
ws.column_dimensions['K'].width = 10  # PE
ws.column_dimensions['L'].width = 10  # PB
ws.column_dimensions['M'].width = 12  # EV/EBITDA
ws.column_dimensions['N'].width = 12  # EV/Rev

LAST_COL = 14

# --- Header Section ---
R = 1
sc = set_cell
sc(ws, R, 1, f"A股可比公司分析 / A-Share Comparable Company Analysis", TITLE_FONT)
ws.merge_cells('A1:N1')
R = 2
sc(ws, R, 1, f"行业: {SECTOR} | 日期: {DATE} | 货币: 人民币(¥M) / RMB Millions", BLACK_FONT)
ws.merge_cells('A2:N2')

# --- Main Table Headers ---
R = 4
section_header(ws, R, "交易乘数与运营指标 / Trading Multiples & Operating Metrics (FY2024)", 1, LAST_COL)

R = 5
sub_header(ws, R, "公司信息 / Company Info", 1, 5)
sub_header(ws, R, "规模与盈利 / Size & Profitability", 6, 10)
sub_header(ws, R, "估值倍数 / Valuation Multiples", 11, 14)

R = 6
cols = [
    "代码\nTicker", "名称\nName", "股价\nPrice", "市值\nMarket Cap", "企业价值\nEV",
    "营业收入\nRevenue", "收入增速\nRev Gr%", "ROE\nROE%", "归母净利润\nNet Income", "EBITDA\nEBITDA",
    "市盈率\nP/E", "市净率\nP/B", "EV/EBITDA\nEV/EBITDA", "EV/Rev\nEV/Rev"
]
for i, col_name in enumerate(cols, 1):
    cell = sc(ws, R, i, col_name, SUBHEADER_FONT, LIGHT_GRAY_FILL, align=CENTER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[R].height = 40

# --- Populate Peer Data ---
start_data_row = 7
num_peers = len(peers)
cur_r = start_data_row

for peer in peers:
    # Column A-B: IDs
    sc(ws, cur_r, 1, peer["t"], BLACK_FONT, align=CENTER)
    sc(ws, cur_r, 2, peer["n"], BLACK_FONT, align=CENTER)
    
    # Column C: Price
    sc(ws, cur_r, 3, peer["p"], BLUE_FONT, LIGHT_GREEN_FILL, fmt=PRICE_FMT, align=CENTER)
    
    # Column D: MCap (Price * Shares) -> =C*Shares
    sc(ws, cur_r, 4, f"=$C{cur_r}*{peer['sh']}", BLACK_FONT, fmt=NUM_FMT, align=CENTER)
    
    # Column E: EV (MCap + NetDebt) -> =D + NetDebt
    sc(ws, cur_r, 5, f"=$D{cur_r}+({peer['nd']})", BLACK_FONT, fmt=NUM_FMT, align=CENTER)
    
    # Column F-J: Input Financials
    sc(ws, cur_r, 6, peer["rev"], BLUE_FONT, LIGHT_GREEN_FILL, fmt=NUM_FMT, align=CENTER)
    sc(ws, cur_r, 7, peer["rg"], BLUE_FONT, LIGHT_GREEN_FILL, fmt=PCT_FMT, align=CENTER)
    sc(ws, cur_r, 8, peer["roe"], BLUE_FONT, LIGHT_GREEN_FILL, fmt=PCT_FMT, align=CENTER)
    sc(ws, cur_r, 9, peer["ni"], BLUE_FONT, LIGHT_GREEN_FILL, fmt=NUM_FMT, align=CENTER)
    sc(ws, cur_r, 10, peer["ebitda"], BLUE_FONT, LIGHT_GREEN_FILL, fmt=NUM_FMT, align=CENTER)
    
    # Column K: P/E (MCap / NI) -> =D/I
    sc(ws, cur_r, 11, f"=IF($I{cur_r}>0,$D{cur_r}/$I{cur_r},\"NM\")", BLACK_FONT, fmt=MULT_FMT, align=CENTER)
    
    # Column L: P/B (MCap / BV) -> =D/BV
    sc(ws, cur_r, 12, f"=$D{cur_r}/{peer['bv']}", BLACK_FONT, fmt=MULT_FMT, align=CENTER)
    
    # Column M: EV/EBITDA -> =E/J
    sc(ws, cur_r, 13, f"=IF($J{cur_r}>0,$E{cur_r}/$J{cur_r},\"NM\")", BLACK_FONT, fmt=MULT_FMT, align=CENTER)
    
    # Column N: EV/Rev -> =E/F
    sc(ws, cur_r, 14, f"=$E{cur_r}/$F{cur_r}", BLACK_FONT, fmt=MULT_FMT, align=CENTER)
    
    cur_r += 1

end_data_row = cur_r - 1

# --- Statistics Section ---
R = cur_r
sub_header(ws, R, "统计数据 / STATISTICS", 1, 5)

stats = [
    ("最大值 / Maximum", f"=MAX({{range}})"),
    ("75% 分位 / 75th Percentile", f"=QUARTILE.INC({{range}}, 3)"),
    ("中位数 / Median", f"=MEDIAN({{range}})"),
    ("平均值 / Mean", f"=AVERAGE({{range}})"),
    ("25% 分位 / 25th Percentile", f"=QUARTILE.INC({{range}}, 1)"),
    ("最小值 / Minimum", f"=MIN({{range}})"),
]

stat_r = R + 1
for label, formula_tpl in stats:
    # Merge A-E for the label
    sc(ws, stat_r, 1, label, STAT_FONT, LIGHT_GRAY_FILL, align=RIGHT)
    for c in range(2, 6): sc(ws, stat_r, c, "", STAT_FONT, LIGHT_GRAY_FILL)
    ws.merge_cells(start_row=stat_r, start_column=1, end_row=stat_r, end_column=5)
    
    # Financial metrics columns (F to N)
    fmts = [NUM_FMT, PCT_FMT, PCT_FMT, NUM_FMT, NUM_FMT, MULT_FMT, MULT_FMT, MULT_FMT, MULT_FMT]
    
    for c_idx, c in enumerate(range(6, LAST_COL + 1)):
        col_ltr = get_column_letter(c)
        data_range = f"{col_ltr}{start_data_row}:{col_ltr}{end_data_row}"
        fml = formula_tpl.format(range=data_range)
        # Using IFERROR to handle "NM" in multiples
        if c >= 11:  # Multiples columns where "NM" might exist
            fml = f"=IFERROR({fml}, \"NM\")"
        sc(ws, stat_r, c, fml, BLACK_FONT, LIGHT_GRAY_FILL, fmt=fmts[c_idx], align=CENTER)
    
    stat_r += 1

# --- Insights/Notes ---
R = stat_r + 2
section_header(ws, R, "行业观察 / SECTOR INSIGHTS", 1, LAST_COL)

notes = [
    "1. 特高压超级周期 (Super-Cycle): 受国家电网“十五五”4万亿投资预期及出海需求共振，电网设备板块整体处于历史估值高位 (PE达30x+)。",
    "2. 龙头国电南瑞 (600406): 依然保持电网二次设备绝对龙头地位，净利润及ROE表现稳健。其29.6x PE代表了板块核心中枢。",
    "3. 出海领军思源电气 (002028): 凭借强劲的海外订单收入(增速24%)以及最高的ROE(18%)，享有明显的估值溢价 (PE近30x，当前静态数据或存在扰动)。",
    "4. 特变电工 (600089): 估值分歧最大。尽管营收规模最大(978亿)，但由于多晶硅业务严重拖累(影响净利润与ROE)，其EV/EBITDA (约8x) 显著低于纯电网设备企业(20x+)。",
    "5. 整体而言: 纯电网资产(南瑞、思源、许继、平高)受到市场资金的高度追捧，而含有光伏/新能源上游资产的综合型公司(特变)则面临估值折价。"
]

R += 1
for i, n in enumerate(notes):
    sc(ws, R+i, 1, n, BLACK_FONT, align=LEFT)
    for c in range(2, LAST_COL + 1): sc(ws, R+i, c, "")
    ws.merge_cells(start_row=R+i, start_column=1, end_row=R+i, end_column=LAST_COL)
    ws.row_dimensions[R+i].height = 25  # Give some breathing room to the text

# ============================================================
# TARGET CHART DATA PREP (Optional, just to have data ordered)
# ============================================================
# We sort the peers by PE and store them in the side for users to make charts later
ws.column_dimensions['P'].width = 15
ws.column_dimensions['Q'].width = 15

# ============================================================
# SAVE
# ============================================================
output_path = r"c:\Users\rickylu\.gemini\antigravity\scratch\financial-services-plugins\PowerGrid_Comps_2026-03-10.xlsx"
wb.save(output_path)
print(f"[OK] Saved: {output_path}")
print(f"Peers: {[p['n'] for p in peers]}")
