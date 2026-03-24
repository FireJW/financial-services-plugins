"""
磷化工 (Phosphate Chemical Sector) Comps Analysis Model
申万二级：基础化工-磷化工及磷酸盐 / A-Share Sector: Phosphate Chemicals
Generated: 2026-03-11
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# Color & Style Definitions
# ============================================================
BLUE_FONT = Font(name="Microsoft YaHei", size=11, color="0000FF")       
BLACK_FONT = Font(name="Microsoft YaHei", size=11, color="000000")      
GREEN_FONT = Font(name="Microsoft YaHei", size=11, color="008000")      

HEADER_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="000000")
TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color="000000")
STAT_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="000000", italic=True)

DARK_BLUE = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

PCT_FMT = '0.0%'
NUM_FMT = '#,##0'
MULT_FMT = '0.0"x"'
PRICE_FMT = '¥#,##0.00'

def sc(ws, r, c, v, f=None, fl=None, fm=None, al=None):
    cell = ws.cell(row=r, column=c, value=v)
    if f: cell.font = f
    if fl: cell.fill = fl
    if fm: cell.number_format = fm
    if al: cell.alignment = al
    cell.border = THIN_BORDER
    return cell

def hdr(ws, r, txt, c1=1, c2=16):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c)
        cell.fill = DARK_BLUE; cell.font = HEADER_FONT; cell.alignment = CENTER; cell.border = THIN_BORDER
    ws.cell(row=r, column=c1, value=txt)
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

def shdr(ws, r, txt, c1=1, c2=16):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c)
        cell.fill = LIGHT_BLUE; cell.font = SUBHEADER_FONT; cell.alignment = CENTER; cell.border = THIN_BORDER
    ws.cell(row=r, column=c1, value=txt)

# ============================================================
# TARGET & PEER DATA (LTM / FY2024A)
# RMB in Millions (Except per share)
# ============================================================
DATE = "2026-03-11"
SECTOR = "磷化工 (Phosphate Chemical Sector)"

# Data based on FY2024 annual reports and recent March 2026 market data
# Ticker, Name, Price(¥), Shares(M), NetDebt(M), Rev(M), NI(M), EBITDA(M), Equity(BV,M), RevGr(YoY), ROE, GrossMargin
peers = [
    {
        "t": "600096.SH", "n": "云天化",
        "p": 41.15, "sh": 1834.3, "nd": 8906,      # nd = 有息负债158.43亿 - 现金69.37亿 = 89.06亿
        "rev": 61537, "ni": 5333, "ebitda": 9854,
        "bv": 22360, "rg": -0.021, "roe": 0.2385, "gm": 0.175
    },
    {
        "t": "600141.SH", "n": "兴发集团",
        "p": 40.55, "sh": 1103.3, "nd": 15800,      # nd ≈ 有息负债244亿 - 现金86亿 ≈ 158亿
        "rev": 28396, "ni": 1619, "ebitda": 4770,
        "bv": 21463, "rg": 0.038, "roe": 0.0603, "gm": 0.168
    },
    {
        "t": "000422.SZ", "n": "湖北宜化",
        "p": 17.55, "sh": 1084.0, "nd": 8900,       # nd ≈ 有息负债126亿 - 现金37亿 ≈ 89亿
        "rev": 16964, "ni": 653, "ebitda": 1900,
        "bv": 8800, "rg": -0.005, "roe": 0.0886, "gm": 0.137
    },
    {
        "t": "002895.SZ", "n": "川恒股份",
        "p": 43.86, "sh": 542.0, "nd": 1705,        # nd ≈ 有息负债27.05亿 - 现金10亿 ≈ 17.05亿
        "rev": 5906, "ni": 956, "ebitda": 1680,
        "bv": 5980, "rg": 0.367, "roe": 0.1603, "gm": 0.331
    },
    {
        "t": "002312.SZ", "n": "川发龙蟒",
        "p": 12.88, "sh": 1889.0, "nd": 4200,       # nd ≈ 估计有息负债72亿 - 现金30亿 ≈ 42亿
        "rev": 8178, "ni": 533, "ebitda": 1450,
        "bv": 9656, "rg": 0.061, "roe": 0.0564, "gm": 0.143
    },
    {
        "t": "000902.SZ", "n": "新洋丰",
        "p": 18.14, "sh": 1255.0, "nd": 2850,       # nd ≈ 估计有息负债40亿 - 现金12亿 ≈ 28.5亿
        "rev": 15563, "ni": 1315, "ebitda": 2480,
        "bv": 10500, "rg": 0.031, "roe": 0.131, "gm": 0.161
    },
]

# ============================================================
# COMPS SHEET
# ============================================================
ws = wb.active
ws.title = "Comps"
ws.sheet_properties.tabColor = "17365D"

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 13
ws.column_dimensions['E'].width = 13
ws.column_dimensions['F'].width = 13
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 12
ws.column_dimensions['L'].width = 10
ws.column_dimensions['M'].width = 10
ws.column_dimensions['N'].width = 12
ws.column_dimensions['O'].width = 12
ws.column_dimensions['P'].width = 10

LAST = 16

# --- Header Section ---
R = 1
sc(ws, R, 1, f"A股可比公司分析 / A-Share Comparable Company Analysis", TITLE_FONT)
ws.merge_cells(f'A1:{get_column_letter(LAST)}1')
R = 2
sc(ws, R, 1, f"行业: {SECTOR} | 日期: {DATE} | 货币: 人民币(¥M) / RMB Millions", BLACK_FONT)
ws.merge_cells(f'A2:{get_column_letter(LAST)}2')

# --- Main Table Headers ---
R = 4
hdr(ws, R, "交易乘数与运营指标 / Trading Multiples & Operating Metrics", 1, LAST)

R = 5
shdr(ws, R, "公司信息 / Company", 1, 5)
shdr(ws, R, "规模与盈利 / Size & Profitability", 6, 12)
shdr(ws, R, "估值倍数 / Valuation", 13, 16)

R = 6
cols = [
    "代码\nTicker", "名称\nName", "股价\nPrice", "市值\nMarket Cap", "企业价值\nEV",
    "营业收入\nRevenue", "收入增速\nRev Gr%", "毛利率\nGross%", "ROE\nROE%",
    "归母净利\nNet Income", "EBITDA\nEBITDA", "净利率\nNet Mgn%",
    "市盈率\nP/E", "市净率\nP/B", "EV/EBITDA\nEV/EBITDA", "EV/Rev\nEV/Rev"
]
for i, name in enumerate(cols, 1):
    c = sc(ws, R, i, name, SUBHEADER_FONT, LIGHT_GRAY, al=CENTER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[R].height = 40

# --- Populate Data ---
sr = 7
cr = sr

for p in peers:
    sc(ws, cr, 1, p["t"], BLACK_FONT, al=CENTER)
    sc(ws, cr, 2, p["n"], BLACK_FONT, al=CENTER)
    sc(ws, cr, 3, p["p"], BLUE_FONT, LIGHT_GREEN, PRICE_FMT, CENTER)
    # Market Cap = Price × Shares(M)
    sc(ws, cr, 4, f"=$C{cr}*{p['sh']}", BLACK_FONT, fm=NUM_FMT, al=CENTER)
    # EV = Market Cap + Net Debt
    sc(ws, cr, 5, f"=$D{cr}+({p['nd']})", BLACK_FONT, fm=NUM_FMT, al=CENTER)
    
    sc(ws, cr, 6, p["rev"], BLUE_FONT, LIGHT_GREEN, NUM_FMT, CENTER)
    sc(ws, cr, 7, p["rg"], BLUE_FONT, LIGHT_GREEN, PCT_FMT, CENTER)
    sc(ws, cr, 8, p["gm"], BLUE_FONT, LIGHT_GREEN, PCT_FMT, CENTER)
    sc(ws, cr, 9, p["roe"], BLUE_FONT, LIGHT_GREEN, PCT_FMT, CENTER)
    sc(ws, cr, 10, p["ni"], BLUE_FONT, LIGHT_GREEN, NUM_FMT, CENTER)
    sc(ws, cr, 11, p["ebitda"], BLUE_FONT, LIGHT_GREEN, NUM_FMT, CENTER)
    # Net Margin = NI / Rev
    sc(ws, cr, 12, f"=$J{cr}/$F{cr}", BLACK_FONT, fm=PCT_FMT, al=CENTER)
    
    # P/E = Market Cap / Net Income
    sc(ws, cr, 13, f'=IF($J{cr}>0,$D{cr}/$J{cr},"NM")', BLACK_FONT, fm=MULT_FMT, al=CENTER)
    # P/B = Market Cap / Book Value
    sc(ws, cr, 14, f"=$D{cr}/{p['bv']}", BLACK_FONT, fm=MULT_FMT, al=CENTER)
    # EV/EBITDA
    sc(ws, cr, 15, f'=IF($K{cr}>0,$E{cr}/$K{cr},"NM")', BLACK_FONT, fm=MULT_FMT, al=CENTER)
    # EV/Revenue
    sc(ws, cr, 16, f"=$E{cr}/$F{cr}", BLACK_FONT, fm=MULT_FMT, al=CENTER)
    cr += 1

er = cr - 1

# --- Statistics ---
R = cr
shdr(ws, R, "统计数据 / STATISTICS", 1, 5)

stats = [
    ("最大值 / Maximum", "=MAX({range})"),
    ("75% 分位 / 75th Percentile", "=QUARTILE.INC({range}, 3)"),
    ("中位数 / Median", "=MEDIAN({range})"),
    ("平均值 / Mean", "=AVERAGE({range})"),
    ("25% 分位 / 25th Percentile", "=QUARTILE.INC({range}, 1)"),
    ("最小值 / Minimum", "=MIN({range})"),
]

str_r = R + 1
for lb, fml_tpl in stats:
    sc(ws, str_r, 1, lb, STAT_FONT, LIGHT_GRAY, al=RIGHT)
    for c in range(2, 6): sc(ws, str_r, c, "", STAT_FONT, LIGHT_GRAY)
    ws.merge_cells(start_row=str_r, start_column=1, end_row=str_r, end_column=5)
    
    # Stats for cols 6-16 (Rev not stat'd; Gr%, GM%, ROE, NI-skip, EBITDA-skip, NetMgn%, PE, PB, EV/EBITDA, EV/Rev)
    # Skip absolute amounts (Rev col6, NI col10, EBITDA col11) - no stats
    # Stats for: col7(RevGr), col8(GM), col9(ROE), col12(NetMgn), col13(PE), col14(PB), col15(EV/EBITDA), col16(EV/Rev)
    stat_cols = [7, 8, 9, 12, 13, 14, 15, 16]
    stat_fmts = {7: PCT_FMT, 8: PCT_FMT, 9: PCT_FMT, 12: PCT_FMT,
                 13: MULT_FMT, 14: MULT_FMT, 15: MULT_FMT, 16: MULT_FMT}
    
    # Fill non-stat columns with blank
    for c in [6, 10, 11]:
        sc(ws, str_r, c, "", BLACK_FONT, LIGHT_GRAY, al=CENTER)
    
    for c in stat_cols:
        ltr = get_column_letter(c)
        fml = fml_tpl.format(range=f"{ltr}{sr}:{ltr}{er}")
        if c >= 13:
            fml_val = fml  # already has = prefix
            fml = f'=IFERROR({fml_val.lstrip("=")}, "NM")'
        sc(ws, str_r, c, fml, BLACK_FONT, LIGHT_GRAY, stat_fmts[c], CENTER)
    str_r += 1

# --- Insights ---
R = str_r + 2
hdr(ws, R, "行业观察 / SECTOR INSIGHTS", 1, LAST)

notes = [
    '1. 行业格局 (Industry Landscape): 磷化工行业正从传统磷肥向「肥料+精细化工+新能源材料」三足鼎立格局转型。磷矿石作为战略性稀缺资源，供需偏紧格局持续，拥有上游磷矿资源的企业享有显著竞争壁垒。',
    '2. 云天化 (600096): 行业绝对龙头，国内唯一磷矿100%自给企业。FY2024营收615亿、归母净利53亿，ROE高达23.9%。已布局磷酸铁产能50万吨(2026E)及固态电池磷材料。P/E约14x，估值显著低于行业均值，反映市场对磷肥周期波动担忧。',
    '3. 兴发集团 (600141): 以「矿电磷硅氟」一体化为核心，全国精细磷产品种类最全企业。电子级磷酸占据领先市场份额，与万华化学合资进军新能源。有息负债偏高(~244亿)拉高EV但体现扩张意图。P/E约28x高于同行，反映精细化工+电子化学品溢价。',
    '4. 川恒股份 (002895): 矿化一体化模式叠加宁德时代深度绑定，FY2024营收同比+36.7%为板块最快增速。ROE 16%、毛利率33%均居行业领先，属磷化工新能源弹性最强标的。高估值(P/E~25x)体现市场对其成长性的认可。',
    '5. 新洋丰 (000902): 国内复合肥渠道与销量龙头，经营稳健。毛利率16%、ROE 13%表现均衡，P/E约17x在板块内估值偏低，适合稳健型投资者。磷矿资源布局渐进但规模仍弱于第一梯队。',
    '6. 湖北宜化 (000422) & 川发龙蟒 (002312): 两者均属第二/三梯队磷化工企业。湖北宜化聚焦化肥+PVC+烧碱，ROE 8.9%回报偏低。川发龙蟒为全球工业级磷酸一铵领军者，ROE 5.6%尚处产能爬坡期。二者P/E分别约29x和46x，隐含市场对磷酸铁锂转型的远期预期。',
]

R += 1
for i, n in enumerate(notes):
    sc(ws, R+i, 1, n, BLACK_FONT, al=LEFT)
    for c in range(2, LAST + 1): sc(ws, R+i, c, "")
    ws.merge_cells(start_row=R+i, start_column=1, end_row=R+i, end_column=LAST)
    ws.row_dimensions[R+i].height = 25

# ============================================================
# SAVE
# ============================================================
output_path = r"c:\Users\rickylu\.gemini\antigravity\scratch\financial-services-plugins\Phosphate_Comps_2026-03-11.xlsx"
wb.save(output_path)
print(f"[OK] Saved: {output_path}")
