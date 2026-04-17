# -*- coding: utf-8 -*-
'''
NVIDIA Liquid Cooling Supply Chain Comps Analysis
Data Center Thermal Management / AI Liquid Cooling
Generated: 2026-03-16
'''

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

BF = Font(name="Microsoft YaHei", size=11, color="0000FF")
KF = Font(name="Microsoft YaHei", size=11, color="000000")
HF = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
SF = Font(name="Microsoft YaHei", size=11, bold=True, color="000000")
TF = Font(name="Microsoft YaHei", size=14, bold=True, color="000000")
STF = Font(name="Microsoft YaHei", size=11, bold=True, color="000000", italic=True)

DB = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
LB = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
LG = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

CA = Alignment(horizontal="center", vertical="center")
LA = Alignment(horizontal="left", vertical="center")
RA = Alignment(horizontal="right", vertical="center")
BD = Border(left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))

P='0.0%'; N='#,##0'; MF='0.0"x"'; PF='#,##0.00'

def sc(ws, r, c, v, f=None, fl=None, fm=None, al=None):
    cell = ws.cell(row=r, column=c, value=v)
    if f: cell.font = f
    if fl: cell.fill = fl
    if fm: cell.number_format = fm
    if al: cell.alignment = al
    else: cell.alignment = CA
    cell.border = BD
    return cell

def hdr(ws, r, txt, c1=1, c2=16):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c)
        cell.fill = DB; cell.font = HF; cell.alignment = CA; cell.border = BD
    ws.cell(row=r, column=c1, value=txt)
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

def shdr(ws, r, txt, c1=1, c2=16):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c)
        cell.fill = LB; cell.font = SF; cell.alignment = CA; cell.border = BD
    ws.cell(row=r, column=c1, value=txt)

DATE = "2026-03-16"
SECTOR = "NVIDIA Liquid Cooling Supply Chain"
LAST = 16

# FY2024 data, prices as of 2026-03-16
# t=ticker, n=name, p=price, sh=shares(M), nd=net_debt(M), rev=revenue(M), ni=net_income(M),
# ebitda=ebitda(M), bv=book_value(M), rg=rev_growth, roe=ROE, gm=gross_margin
peers = [
    {"t": "002837.SZ", "n": "Envicool", "p": 105.00, "sh": 977.0, "nd": 800,
     "rev": 4589, "ni": 453, "ebitda": 741, "bv": 2916, "rg": 0.300, "roe": 0.155, "gm": 0.310},
    {"t": "300499.SZ", "n": "Gaolan", "p": 41.79, "sh": 305.0, "nd": 600,
     "rev": 691, "ni": -50, "ebitda": 35, "bv": 1100, "rg": 0.120, "roe": -0.045, "gm": 0.250},
    {"t": "301018.SZ", "n": "Shenling", "p": 97.14, "sh": 266.0, "nd": 1200,
     "rev": 3016, "ni": 116, "ebitda": 310, "bv": 2500, "rg": 0.201, "roe": 0.047, "gm": 0.240},
    {"t": "300990.SZ", "n": "Tongfei", "p": 88.10, "sh": 171.0, "nd": -200,
     "rev": 2160, "ni": 153, "ebitda": 350, "bv": 1890, "rg": 0.171, "roe": 0.081, "gm": 0.221},
    {"t": "872808.BJ", "n": "Sugon DC", "p": 90.39, "sh": 200.0, "nd": -150,
     "rev": 506, "ni": 61, "ebitda": 100, "bv": 700, "rg": -0.222, "roe": 0.087, "gm": 0.301},
    {"t": "002126.SZ", "n": "Yinlun", "p": 45.21, "sh": 846.0, "nd": 2500,
     "rev": 12702, "ni": 784, "ebitda": 1650, "bv": 6057, "rg": 0.153, "roe": 0.129, "gm": 0.201},
]

ws = wb.active; ws.title = "Comps"; ws.sheet_properties.tabColor = "17365D"
for i, w in enumerate([12,12,10,13,13,13,10,10,10,12,12,10,10,12,12,10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

sc(ws, 1, 1, "A-Share Comparable Company Analysis / NVIDIA Liquid Cooling Supply Chain", TF)
ws.merge_cells(f'A1:{get_column_letter(LAST)}1')
sc(ws, 2, 1, f"Sector: {SECTOR} | Date: {DATE} | Unit: RMB Millions (M)", KF)
ws.merge_cells(f'A2:{get_column_letter(LAST)}2')

R = 4
hdr(ws, R, "Trading Multiples & Operating Metrics", 1, LAST)
R = 5
shdr(ws, R, "Company", 1, 5)
shdr(ws, R, "Size & Profitability", 6, 12)
shdr(ws, R, "Valuation", 13, 16)

R = 6
cols = ["Ticker\nTicker","Name\nName","Price\nPrice","MCap\nMarket Cap","EV\nEV",
        "Revenue\nRevenue","Rev Gr%\nRev Gr%","Gross%\nGross%","ROE\nROE%",
        "Net Inc\nNet Income","EBITDA\nEBITDA","Net Mgn%\nNet Mgn%",
        "P/E\nP/E","P/B\nP/B","EV/EBITDA\nEV/EBITDA","EV/Rev\nEV/Rev"]
for i, name in enumerate(cols, 1):
    c = sc(ws, R, i, name, SF, GY, al=CA)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[R].height = 40

sr = 7; cr = sr
for p in peers:
    sc(ws, cr, 1, p["t"], KF, al=CA)
    sc(ws, cr, 2, p["n"], KF, al=CA)
    sc(ws, cr, 3, p["p"], BF, LG, PF, CA)
    sc(ws, cr, 4, f"=$C{cr}*{p['sh']}", KF, fm=N, al=CA)
    sc(ws, cr, 5, f"=$D{cr}+({p['nd']})", KF, fm=N, al=CA)
    sc(ws, cr, 6, p["rev"], BF, LG, N, CA)
    sc(ws, cr, 7, p["rg"], BF, LG, P, CA)
    sc(ws, cr, 8, p["gm"], BF, LG, P, CA)
    sc(ws, cr, 9, p["roe"], BF, LG, P, CA)
    sc(ws, cr, 10, p["ni"], BF, LG, N, CA)
    sc(ws, cr, 11, p["ebitda"], BF, LG, N, CA)
    sc(ws, cr, 12, f"=$J{cr}/$F{cr}", KF, fm=P, al=CA)
    sc(ws, cr, 13, f'=IF($J{cr}>0,$D{cr}/$J{cr},"NM")', KF, fm=MF, al=CA)
    sc(ws, cr, 14, f"=$D{cr}/{p['bv']}", KF, fm=MF, al=CA)
    sc(ws, cr, 15, f'=IF($K{cr}>0,$E{cr}/$K{cr},"NM")', KF, fm=MF, al=CA)
    sc(ws, cr, 16, f"=$E{cr}/$F{cr}", KF, fm=MF, al=CA)
    cr += 1
er = cr - 1

R = cr
shdr(ws, R, "STATISTICS", 1, 5)
stats = [("Max / Maximum","=MAX({range})"),("75th / 75th Pctile","=QUARTILE.INC({range}, 3)"),
         ("Median / Median","=MEDIAN({range})"),("Mean / Mean","=AVERAGE({range})"),
         ("25th / 25th Pctile","=QUARTILE.INC({range}, 1)"),("Min / Minimum","=MIN({range})")]
str_r = R + 1
stat_cols = [7,8,9,12,13,14,15,16]
stat_fmts = {7:P,8:P,9:P,12:P,13:MF,14:MF,15:MF,16:MF}
for lb, fml_tpl in stats:
    sc(ws, str_r, 1, lb, STF, GY, al=RA)
    for c in range(2, 6): sc(ws, str_r, c, "", STF, GY)
    ws.merge_cells(start_row=str_r, start_column=1, end_row=str_r, end_column=5)
    for c in [6, 10, 11]: sc(ws, str_r, c, "", KF, GY, al=CA)
    for c in stat_cols:
        ltr = get_column_letter(c)
        fml = fml_tpl.format(range=f"{ltr}{sr}:{ltr}{er}")
        if c >= 13: fml = f'=IFERROR({fml.lstrip("=")}, "NM")'
        sc(ws, str_r, c, fml, KF, GY, stat_fmts[c], CA)
    str_r += 1

R = str_r + 2
hdr(ws, R, "SECTOR INSIGHTS", 1, LAST)
notes = [
    '1. NVIDIA Liquid Cooling Megatrend: GB300/Rubin chips mandate 100% liquid cooling by 2026. DC liquid cooling penetration expected to surge from 14%(2024) to 33%(2025) and higher. China liquid cooling server market CAGR 48% (2025-2029).',
    '2. Envicool (002837): NVIDIA core supplier - cold plates, UQD quick-connects, CDUs. GB300 cabinet value share 35% (highest among China suppliers). FY2024 rev +30%, liquid cooling revenue ~2x YoY. 1200-unit GB300 order worth RMB 9.6-10.2B. P/E ~226x reflects extreme growth premium.',
    '3. Gaolan (300499): NVIDIA GB200/GB300 certified. Immersion cooling for ByteDance Singapore AI DC. FY2024 net loss, turnaround play. Expected ~2B incremental revenue from DGX B100 liquid cooling plates in 2026.',
    '4. Shenling (301018): Liquid cooling system revenue +190% YoY in H1 2025. Deep partnership with Huawei. Cold plate solutions compatible with GB300 high-power chips. P/E ~223x, growth-stage valuation.',
    '5. Tongfei (300990): Full liquid cooling solutions (cold plate + immersion). Net cash position. Most reasonable valuation in the peer group at P/E ~98x. FY2024 rev +17%, steady grower.',
    '6. Yinlun (002126): Diversified thermal management (auto + DC). GB200/GB300 micro-channel cold plate core supplier. FY2024 rev 12.7B (+15%), P/E ~49x, cheapest in group due to auto exposure dilution.',
]
R += 1
for i, n in enumerate(notes):
    sc(ws, R+i, 1, n, KF, al=LA)
    for c in range(2, LAST + 1): sc(ws, R+i, c, "")
    ws.merge_cells(start_row=R+i, start_column=1, end_row=R+i, end_column=LAST)
    ws.row_dimensions[R+i].height = 28

out = r"c:\Users\rickylu\.gemini\antigravity\scratch\financial-services-plugins\LiquidCooling_Comps_2026-03-16.xlsx"
wb.save(out)
print(f"[OK] Saved: {out}")
