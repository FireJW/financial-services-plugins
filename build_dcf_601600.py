"""
中国铝业 (601600.SH) DCF Valuation Model + Aluminum Price Sensitivity
Generated: 2026-03-09
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# Styles
# ============================================================
BLUE_FONT = Font(name="Microsoft YaHei", size=11, color="0000FF")
BLACK_FONT = Font(name="Microsoft YaHei", size=11, color="000000")
GREEN_FONT = Font(name="Microsoft YaHei", size=11, color="008000")
HEADER_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="000000")
TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color="000000")
RESULT_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="000000")
RED_BOLD = Font(name="Microsoft YaHei", size=14, bold=True, color="FF0000")
RED_12 = Font(name="Microsoft YaHei", size=12, bold=True, color="FF0000")

DARK_BLUE = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

C = Alignment(horizontal="center", vertical="center")
BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))

P = '0.0%'
N = '#,##0'
N1 = '#,##0.0'
PR = '#,##0.00'

def sc(ws, r, c, v, font=None, fill=None, fmt=None, cmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font: cell.font = font
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    cell.alignment = C
    cell.border = BORDER
    if cmt:
        cell.comment = openpyxl.comments.Comment(cmt, "DCF")
    return cell

def hdr(ws, r, txt, c1=1, c2=10):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c)
        cell.fill = DARK_BLUE; cell.font = HEADER_FONT; cell.alignment = C; cell.border = BORDER
    ws.cell(row=r, column=c1, value=txt)
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

def shdr(ws, r, txt, c1=1, c2=10):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c)
        cell.fill = LIGHT_BLUE; cell.font = SUBHEADER_FONT; cell.alignment = C; cell.border = BORDER
    ws.cell(row=r, column=c1, value=txt)

# ============================================================
# DATA
# ============================================================
COMPANY = "Chalco"
TICKER = "601600.SH"
DATE = "2026-03-09"

PRICE = 14.10
SHARES = 17155     # M shares (171.55 yi)
MCAP = PRICE * SHARES

hist = {
    "year":    [2021,     2022,     2023,     2024],
    "rev":     [270200,   290988,   225071,   237066],  # M RMB
    "gp":      [29700,    33400,    29000,    37337],
    "gp_pct":  [0.110,    0.115,    0.129,    0.1575],
    "ebit":    [18500,    20500,    18000,    30500],
    "ebit_pct":[0.068,    0.070,    0.080,    0.1287],
    "tax":     [3700,     4100,     3600,     6100],
    "ni":      [10980,    4192,     6717,     12400],
    "da":      [12000,    13000,    13500,    14000],
    "capex":   [8000,     9000,     9500,     10000],
    "ebitda":  [30500,    33500,    31500,    44500],
}

DEBT = 85000       # M estimated interest-bearing debt
CASH = 22210       # M cash
NET_DEBT = DEBT - CASH  # 62,790 M

RF = 0.022
BETA = 1.58
ERP = 0.065
KE = RF + BETA * ERP
KD_PRE = 0.038
TAX = 0.20
KD_POST = KD_PRE * (1 - TAX)
EV = MCAP + NET_DEBT
EQ_W = MCAP / EV
DT_W = NET_DEBT / EV
WACC = KE * EQ_W + KD_POST * DT_W

proj = {
    "years": [2025, 2026, 2027, 2028, 2029],
    "bear":  {"rg": [0.03, 0.02, 0.02, 0.01, 0.01], "em": [0.11, 0.10, 0.10, 0.09, 0.09]},
    "base":  {"rg": [0.06, 0.05, 0.04, 0.03, 0.03], "em": [0.13, 0.12, 0.12, 0.11, 0.11]},
    "bull":  {"rg": [0.10, 0.08, 0.06, 0.05, 0.04], "em": [0.15, 0.14, 0.14, 0.13, 0.13]},
}
DA_P = 0.058
CX_P = 0.042
NW_P = 0.01
TG = {"bear": 0.02, "base": 0.025, "bull": 0.035}

# Production data
AL_VOL = 761       # 10k tons aluminum
ALO_VOL = 1687     # 10k tons alumina
AL_PRICE_BASE = 20200  # RMB/ton avg realized 2024
AL_REV_PCT = 0.65  # Al products as % of revenue
ALO_CORR = 0.55    # Alumina price correlation to Al

# ============================================================
# DCF SHEET
# ============================================================
ws = wb.active
ws.title = "DCF"
ws.sheet_properties.tabColor = "17365D"
ws.column_dimensions['A'].width = 32
for c in range(2, 12):
    ws.column_dimensions[get_column_letter(c)].width = 16

R = 1
sc(ws, R, 1, f"{COMPANY} ({TICKER}) DCF", TITLE_FONT)
ws.merge_cells('A1:J1')
R = 2
sc(ws, R, 1, f"TICKER: {TICKER} | DATE: {DATE} | FY2024", BLACK_FONT)
ws.merge_cells('A2:J2')
R = 3
sc(ws, R, 1, "Unit: RMB Millions / M", BLACK_FONT)
ws.merge_cells('A3:J3')

# Case Selector
R = 5
hdr(ws, R, "CASE SELECTOR", 1, 10)
R = 6
sc(ws, R, 1, "Active Case (1=Bear, 2=Base, 3=Bull)", SUBHEADER_FONT, YELLOW)
sc(ws, R, 2, 2, BLUE_FONT, YELLOW, cmt="User input: 1/2/3")

# Market Data
R = 8
hdr(ws, R, "MARKET DATA", 1, 10)
mkt = [
    ("Stock Price", PRICE, PR, "Source: 2026-03-09"),
    ("Shares Outstanding (M)", SHARES, N, "Source: 171.55yi shares"),
    ("Market Cap (M)", None, N, None),
    ("Total Debt (M)", DEBT, N, "Source: estimated from leverage"),
    ("Cash (M)", CASH, N, "Source: 2024 annual report"),
    ("Net Debt (M)", None, N, None),
    ("Enterprise Value (M)", None, N, None),
]
for i, (lb, v, f, cm) in enumerate(mkt):
    r = 9 + i
    sc(ws, r, 1, lb, SUBHEADER_FONT, LIGHT_GRAY)
    if v is not None:
        sc(ws, r, 2, v, BLUE_FONT, LIGHT_GREEN, f, cm)
    elif i == 2: sc(ws, r, 2, "=B9*B10", BLACK_FONT, fmt=f)
    elif i == 5: sc(ws, r, 2, "=B12-B13", BLACK_FONT, fmt=f)
    elif i == 6: sc(ws, r, 2, "=B11+B14", BLACK_FONT, fmt=f)

# Scenario Assumptions
R = 17
hdr(ws, R, "SCENARIO ASSUMPTIONS", 1, 10)
hp = ["Assumption"] + [f"FY{y}E" for y in proj["years"]]

def write_case(ws, start_r, name, case_key, tg_key):
    shdr(ws, start_r, name, 1, 10)
    for c, h in enumerate(hp, 1):
        sc(ws, start_r+1, c, h, SUBHEADER_FONT, LIGHT_BLUE)
    r = start_r + 2
    sc(ws, r, 1, "Revenue Growth", SUBHEADER_FONT, LIGHT_GRAY)
    for j, v in enumerate(proj[case_key]["rg"]):
        sc(ws, r, 2+j, v, BLUE_FONT, LIGHT_GREEN, P)
    r += 1
    sc(ws, r, 1, "EBIT Margin", SUBHEADER_FONT, LIGHT_GRAY)
    for j, v in enumerate(proj[case_key]["em"]):
        sc(ws, r, 2+j, v, BLUE_FONT, LIGHT_GREEN, P)
    r += 1
    sc(ws, r, 1, "Terminal Growth", SUBHEADER_FONT, LIGHT_GRAY)
    sc(ws, r, 2, TG[tg_key], BLUE_FONT, LIGHT_GREEN, P)

write_case(ws, 18, "BEAR CASE", "bear", "bear")      # rows 18-22
write_case(ws, 23, "BASE CASE", "base", "base")      # rows 23-27
write_case(ws, 28, "BULL CASE", "bull", "bull")       # rows 28-32

# Selected Case (Row 33-37)
R = 33
shdr(ws, R, "SELECTED CASE (Driven by B6)", 1, 10)
R = 34
for c, h in enumerate(hp, 1):
    sc(ws, R, c, h, SUBHEADER_FONT, LIGHT_BLUE)

R = 35  # Rev Growth
sc(ws, R, 1, "Revenue Growth", SUBHEADER_FONT, YELLOW)
for j in range(5):
    cl = get_column_letter(2+j)
    sc(ws, R, 2+j, f'=IF($B$6=1,{cl}$20,IF($B$6=2,{cl}$25,{cl}$30))', BLACK_FONT, YELLOW, P)

R = 36  # EBIT Margin
sc(ws, R, 1, "EBIT Margin", SUBHEADER_FONT, YELLOW)
for j in range(5):
    cl = get_column_letter(2+j)
    sc(ws, R, 2+j, f'=IF($B$6=1,{cl}$21,IF($B$6=2,{cl}$26,{cl}$31))', BLACK_FONT, YELLOW, P)

R = 37  # Terminal Growth
sc(ws, R, 1, "Terminal Growth", SUBHEADER_FONT, YELLOW)
sc(ws, R, 2, '=IF($B$6=1,$B$22,IF($B$6=2,$B$27,$B$32))', BLACK_FONT, YELLOW, P)

# Common assumptions
R = 38
sc(ws, R, 1, "D&A % Rev", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws, R, 2, DA_P, BLUE_FONT, LIGHT_GREEN, P, "Source: Historical ~5.8%")
R = 39
sc(ws, R, 1, "CapEx % Rev", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws, R, 2, CX_P, BLUE_FONT, LIGHT_GREEN, P, "Source: Historical ~4.2%")
R = 40
sc(ws, R, 1, "NWC % dRev", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws, R, 2, NW_P, BLUE_FONT, LIGHT_GREEN, P)
R = 41
sc(ws, R, 1, "Tax Rate", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws, R, 2, TAX, BLUE_FONT, LIGHT_GREEN, P, "Source: ~20% effective")

# P&L Projection
R = 43
hdr(ws, R, "INCOME STATEMENT & FREE CASH FLOW", 1, 10)
R = 44
yrs = [""] + [f"FY{y}A" for y in hist["year"]] + [f"FY{y}E" for y in proj["years"]]
for c, h in enumerate(yrs, 1):
    fl = LIGHT_BLUE if "A" in str(h) else LIGHT_GRAY if "E" in str(h) else None
    sc(ws, R, c, h, SUBHEADER_FONT, fl)

# Row 45: Revenue
R = 45
sc(ws, R, 1, "Revenue", SUBHEADER_FONT)
for i, rv in enumerate(hist["rev"]):
    sc(ws, R, 2+i, rv, BLUE_FONT, fmt=N, cmt=f"Source: FY{hist['year'][i]}")
for j in range(5):
    pc = get_column_letter(5+j)
    gc = get_column_letter(2+j)
    sc(ws, R, 6+j, f"={pc}45*(1+{gc}$35)", BLACK_FONT, fmt=N)

# Row 46: Growth
R = 46
sc(ws, R, 1, "  Growth %", BLACK_FONT)
for i in range(1, 4):
    sc(ws, R, 2+i, f"={get_column_letter(2+i)}45/{get_column_letter(1+i)}45-1", BLACK_FONT, fmt=P)
for j in range(5):
    sc(ws, R, 6+j, f"={get_column_letter(6+j)}45/{get_column_letter(5+j)}45-1", BLACK_FONT, fmt=P)

# Row 47: EBIT
R = 47
sc(ws, R, 1, "EBIT", SUBHEADER_FONT)
for i, eb in enumerate(hist["ebit"]):
    sc(ws, R, 2+i, eb, BLUE_FONT, fmt=N)
for j in range(5):
    sc(ws, R, 6+j, f"={get_column_letter(6+j)}45*{get_column_letter(2+j)}$36", BLACK_FONT, fmt=N)

# Row 48: EBIT Margin
R = 48
sc(ws, R, 1, "  EBIT Margin", BLACK_FONT)
for i in range(4):
    sc(ws, R, 2+i, f"={get_column_letter(2+i)}47/{get_column_letter(2+i)}45", BLACK_FONT, fmt=P)
for j in range(5):
    sc(ws, R, 6+j, f"={get_column_letter(6+j)}47/{get_column_letter(6+j)}45", BLACK_FONT, fmt=P)

# Row 49: Tax
R = 49
sc(ws, R, 1, "(-) Taxes", BLACK_FONT)
for i, tx in enumerate(hist["tax"]):
    sc(ws, R, 2+i, -abs(tx), BLUE_FONT, fmt=N)
for j in range(5):
    sc(ws, R, 6+j, f"=-{get_column_letter(6+j)}47*$B$41", BLACK_FONT, fmt=N)

# Row 50: NOPAT
R = 50
sc(ws, R, 1, "NOPAT", SUBHEADER_FONT)
for i in range(4):
    sc(ws, R, 2+i, f"={get_column_letter(2+i)}47+{get_column_letter(2+i)}49", BLACK_FONT, fmt=N)
for j in range(5):
    sc(ws, R, 6+j, f"={get_column_letter(6+j)}47+{get_column_letter(6+j)}49", BLACK_FONT, fmt=N)

# Row 51: D&A
R = 51
sc(ws, R, 1, "(+) D&A", BLACK_FONT)
for i, da in enumerate(hist["da"]):
    sc(ws, R, 2+i, da, BLUE_FONT, fmt=N)
for j in range(5):
    sc(ws, R, 6+j, f"={get_column_letter(6+j)}45*$B$38", BLACK_FONT, fmt=N)

# Row 52: CapEx
R = 52
sc(ws, R, 1, "(-) CapEx", BLACK_FONT)
for i, cx in enumerate(hist["capex"]):
    sc(ws, R, 2+i, -abs(cx), BLUE_FONT, fmt=N)
for j in range(5):
    sc(ws, R, 6+j, f"=-{get_column_letter(6+j)}45*$B$39", BLACK_FONT, fmt=N)

# Row 53: NWC
R = 53
sc(ws, R, 1, "(-) dNWC", BLACK_FONT)
for i in range(4):
    sc(ws, R, 2+i, 0, BLUE_FONT, fmt=N)
for j in range(5):
    sc(ws, R, 6+j, f"=-({get_column_letter(6+j)}45-{get_column_letter(5+j)}45)*$B$40", BLACK_FONT, fmt=N)

# Row 54: UFCF
R = 54
sc(ws, R, 1, "Unlevered FCF", SUBHEADER_FONT, LIGHT_GRAY)
for i in range(4):
    cl = get_column_letter(2+i)
    sc(ws, R, 2+i, f"={cl}50+{cl}51+{cl}52+{cl}53", RESULT_FONT, LIGHT_GRAY, N)
for j in range(5):
    cl = get_column_letter(6+j)
    sc(ws, R, 6+j, f"={cl}50+{cl}51+{cl}52+{cl}53", RESULT_FONT, LIGHT_GRAY, N)

# DCF Valuation
R = 56
hdr(ws, R, "DCF VALUATION", 1, 10)
R = 57
sc(ws, R, 1, "WACC", SUBHEADER_FONT, YELLOW)
sc(ws, R, 2, "=WACC!B16", GREEN_FONT, YELLOW, P)

R = 58
sc(ws, R, 1, "Discount Period", SUBHEADER_FONT)
for j in range(5):
    sc(ws, R, 6+j, 0.5+j, BLUE_FONT, fmt=N1)

R = 59
sc(ws, R, 1, "Discount Factor", BLACK_FONT)
for j in range(5):
    sc(ws, R, 6+j, f"=1/(1+$B$57)^{get_column_letter(6+j)}58", BLACK_FONT, fmt='0.0000')

R = 60
sc(ws, R, 1, "PV of FCF", SUBHEADER_FONT)
for j in range(5):
    cl = get_column_letter(6+j)
    sc(ws, R, 6+j, f"={cl}54*{cl}59", BLACK_FONT, fmt=N)

R = 62
sc(ws, R, 1, "Terminal FCF", BLACK_FONT)
sc(ws, R, 2, "=J54*(1+B37)", BLACK_FONT, fmt=N)
R = 63
sc(ws, R, 1, "Terminal Value", BLACK_FONT)
sc(ws, R, 2, "=B62/(B57-B37)", BLACK_FONT, fmt=N)
R = 64
sc(ws, R, 1, "PV of Terminal Value", SUBHEADER_FONT)
sc(ws, R, 2, "=B63/(1+B57)^4.5", BLACK_FONT, fmt=N)

# Summary
R = 66
hdr(ws, R, "VALUATION SUMMARY", 1, 10)
sc(ws, 67, 1, "Sum PV of FCFs", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws, 67, 2, "=SUM(F60:J60)", BLACK_FONT, LIGHT_GRAY, N)
sc(ws, 68, 1, "PV of Terminal Value", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws, 68, 2, "=B64", BLACK_FONT, LIGHT_GRAY, N)
sc(ws, 69, 1, "Enterprise Value", SUBHEADER_FONT, YELLOW)
sc(ws, 69, 2, "=B67+B68", RESULT_FONT, YELLOW, N)
sc(ws, 70, 1, "(-) Net Debt", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws, 70, 2, "=-B14", BLACK_FONT, LIGHT_GRAY, N)
sc(ws, 71, 1, "Equity Value", SUBHEADER_FONT, YELLOW)
sc(ws, 71, 2, "=B69+B70", RESULT_FONT, YELLOW, N)
sc(ws, 72, 1, "Shares (M)", BLACK_FONT)
sc(ws, 72, 2, "=B10", GREEN_FONT, fmt=N)
sc(ws, 73, 1, "IMPLIED PRICE", RED_BOLD, YELLOW)
sc(ws, 73, 2, "=B71/B72", RED_BOLD, YELLOW, PR)
sc(ws, 74, 1, "Current Price", SUBHEADER_FONT)
sc(ws, 74, 2, "=B9", GREEN_FONT, fmt=PR)
sc(ws, 75, 1, "Implied Upside/(Downside)", SUBHEADER_FONT, YELLOW)
sc(ws, 75, 2, "=B73/B74-1", RESULT_FONT, YELLOW, P)

# Sensitivity Tables
R = 77
hdr(ws, R, "SENSITIVITY ANALYSIS", 1, 10)

# Table 1: WACC vs TGR
R = 78
shdr(ws, R, "Table 1: Price vs WACC & Terminal Growth", 1, 8)
R = 79
tgv = [0.015, 0.020, 0.025, 0.030, 0.035]
wv = [0.08, 0.09, 0.10, 0.11, 0.12]
sc(ws, R, 1, "WACC \\ TGR", SUBHEADER_FONT, LIGHT_BLUE)
for j, t in enumerate(tgv):
    sc(ws, R, 2+j, t, SUBHEADER_FONT, LIGHT_BLUE, P)
for i, w in enumerate(wv):
    r = R+1+i
    sc(ws, r, 1, w, BLUE_FONT, fmt=P)
    for j, t in enumerate(tgv):
        pvs = "+".join([f"{get_column_letter(6+k)}54/(1+$A{r})^{0.5+k}" for k in range(5)])
        tc = f"${get_column_letter(2+j)}${R}"
        wc = f"$A{r}"
        tv = f"(J54*(1+{tc})/({wc}-{tc}))/(1+{wc})^4.5"
        sc(ws, r, 2+j, f"=({pvs}+{tv}+B70)/B72", BLACK_FONT, fmt=PR)

# Table 2: Rev Growth vs EBIT Margin
R = 86
shdr(ws, R, "Table 2: Price vs Rev Growth & EBIT Margin", 1, 8)
R = 87
rgv = [0.03, 0.05, 0.06, 0.08, 0.10]
emv = [0.09, 0.11, 0.12, 0.13, 0.15]
sc(ws, R, 1, "RevGr \\ EBIT%", SUBHEADER_FONT, LIGHT_BLUE)
for j, e in enumerate(emv):
    sc(ws, R, 2+j, e, SUBHEADER_FONT, LIGHT_BLUE, P)
for i, rg in enumerate(rgv):
    r = R+1+i
    sc(ws, r, 1, rg, BLUE_FONT, fmt=P)
    for j, e in enumerate(emv):
        ec = f"${get_column_letter(2+j)}${R}"
        rc = f"$A{r}"
        fcf = f"(E45*(1+{rc})*{ec}*(1-$B$41)+E45*(1+{rc})*$B$38-E45*(1+{rc})*$B$39-E45*{rc}*$B$40)"
        ag = f"({rc}/2+$B$37/2)"
        sc(ws, r, 2+j, f"=({fcf}/($B$57-{ag})+B70)/B72", BLACK_FONT, fmt=PR)

# Table 3: Beta vs Rf
R = 94
shdr(ws, R, "Table 3: Price vs Beta & Risk-Free Rate", 1, 8)
R = 95
bv = [1.0, 1.2, 1.4, 1.58, 1.8]
rv = [0.015, 0.020, 0.022, 0.025, 0.030]
sc(ws, R, 1, "Beta \\ Rf", SUBHEADER_FONT, LIGHT_BLUE)
for j, rf in enumerate(rv):
    sc(ws, R, 2+j, rf, SUBHEADER_FONT, LIGHT_BLUE, P)
for i, b in enumerate(bv):
    r = R+1+i
    sc(ws, r, 1, b, BLUE_FONT, fmt='0.00')
    for j, rf in enumerate(rv):
        rc = f"${get_column_letter(2+j)}${R}"
        bc = f"$A{r}"
        ke = f"({rc}+{bc}*{ERP})"
        wn = f"({ke}*{EQ_W:.4f}+{KD_POST:.4f}*{DT_W:.4f})"
        pvs = "+".join([f"{get_column_letter(6+k)}54/(1+{wn})^{0.5+k}" for k in range(5)])
        tv = f"(J54*(1+B37)/({wn}-B37))/(1+{wn})^4.5"
        sc(ws, r, 2+j, f"=({pvs}+{tv}+B70)/B72", BLACK_FONT, fmt=PR)

# ============================================================
# WACC SHEET
# ============================================================
ws2 = wb.create_sheet("WACC")
ws2.sheet_properties.tabColor = "4472C4"
ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 18

hdr(ws2, 1, "WACC CALCULATION", 1, 3)
shdr(ws2, 2, "COST OF EQUITY (CAPM)", 1, 3)
sc(ws2, 3, 1, "Risk-Free Rate (10Y)", BLACK_FONT)
sc(ws2, 3, 2, RF, BLUE_FONT, LIGHT_GREEN, P, "Source: China 10Y bond")
sc(ws2, 4, 1, "Beta (5Y monthly vs CSI300)", BLACK_FONT)
sc(ws2, 4, 2, BETA, BLUE_FONT, LIGHT_GREEN, '0.00', "Source: investing.com")
sc(ws2, 5, 1, "Equity Risk Premium", BLACK_FONT)
sc(ws2, 5, 2, ERP, BLUE_FONT, LIGHT_GREEN, P)
sc(ws2, 6, 1, "Cost of Equity", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws2, 6, 2, "=B3+B4*B5", RESULT_FONT, LIGHT_GRAY, P)

shdr(ws2, 8, "COST OF DEBT", 1, 3)
sc(ws2, 9, 1, "Pre-Tax Cost of Debt", BLACK_FONT)
sc(ws2, 9, 2, KD_PRE, BLUE_FONT, LIGHT_GREEN, P)
sc(ws2, 10, 1, "Tax Rate", BLACK_FONT)
sc(ws2, 10, 2, TAX, BLUE_FONT, LIGHT_GREEN, P)
sc(ws2, 11, 1, "After-Tax Cost of Debt", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws2, 11, 2, "=B9*(1-B10)", RESULT_FONT, LIGHT_GRAY, P)

shdr(ws2, 13, "CAPITAL STRUCTURE", 1, 3)
sc(ws2, 14, 1, "Equity Weight", BLACK_FONT)
sc(ws2, 14, 2, "=DCF!B11/(DCF!B11+DCF!B14)", GREEN_FONT, fmt=P)
sc(ws2, 15, 1, "Debt Weight", BLACK_FONT)
sc(ws2, 15, 2, "=DCF!B14/(DCF!B11+DCF!B14)", GREEN_FONT, fmt=P)
sc(ws2, 16, 1, "WACC", RED_BOLD, YELLOW)
sc(ws2, 16, 2, "=B6*B14+B11*B15", RED_BOLD, YELLOW, P)

# ============================================================
# AL_Price SHEET
# ============================================================
ws3 = wb.create_sheet("AL_Price")
ws3.sheet_properties.tabColor = "C00000"
ws3.column_dimensions['A'].width = 38
for c in range(2, 10):
    ws3.column_dimensions[get_column_letter(c)].width = 18

hdr(ws3, 1, "SHFE AL PRICE SENSITIVITY", 1, 9)
sc(ws3, 2, 1, "Aluminum price impact on revenue, EBIT, FCF and valuation", BLACK_FONT)
ws3.merge_cells('A2:I2')

# Production Volume
hdr(ws3, 4, "PRODUCTION VOLUME (FY2024A)", 1, 9)
vols = [
    ("Al (alloy) production (10k tons)", AL_VOL, N, "Source: 2024 annual 761"),
    ("Alumina production (10k tons)", ALO_VOL, N, "Source: 2024 annual 1687"),
    ("", None, None, None),
    ("FY2024A Avg Realized Al Price (RMB/ton)", AL_PRICE_BASE, N, "Source: ~20200 estimated"),
    ("FY2024A Revenue (M)", 237066, N, "Source: =DCF!E45"),
]
for i, (lb, v, f, cm) in enumerate(vols):
    r = 5 + i
    sc(ws3, r, 1, lb, SUBHEADER_FONT, LIGHT_GRAY)
    if v is not None:
        sc(ws3, r, 2, v, BLUE_FONT, LIGHT_GREEN, f, cm)

# Assumptions
hdr(ws3, 11, "PRICE-REVENUE ASSUMPTIONS", 1, 9)
sc(ws3, 12, 1, "Al products as % of total revenue", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws3, 12, 2, AL_REV_PCT, BLUE_FONT, LIGHT_GREEN, P, "Al alloy ~65% of total rev")
sc(ws3, 13, 1, "Alumina price correlation to Al price", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws3, 13, 2, ALO_CORR, BLUE_FONT, LIGHT_GREEN, P, "~55% correlated")
sc(ws3, 14, 1, "EBIT operating leverage to Al price", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws3, 14, 2, 1.8, BLUE_FONT, LIGHT_GREEN, '0.0', "1.8x: higher leverage than Hongqiao (more integrated)")
sc(ws3, 15, 1, "Organic volume growth (ex-price)", SUBHEADER_FONT, LIGHT_GRAY)
sc(ws3, 15, 2, 0.03, BLUE_FONT, LIGHT_GREEN, P, "3% volume growth")

# Price Impact Table
hdr(ws3, 17, "SHFE AL PRICE -> REVENUE & VALUATION", 1, 9)
alps = [19000, 20000, 21000, 22000, 23000, 24000, 25000, 26000]

R = 18
sc(ws3, R, 1, "SHFE Al Price (RMB/ton)", SUBHEADER_FONT, LIGHT_BLUE)
for j, p in enumerate(alps):
    sc(ws3, R, 2+j, p, SUBHEADER_FONT, LIGHT_BLUE, N)

R = 19  # Price change vs base
sc(ws3, R, 1, "Al Price Change vs FY2024A Base", BLACK_FONT, LIGHT_GRAY)
for j in range(8):
    sc(ws3, R, 2+j, f"={get_column_letter(2+j)}18/$B$8-1", BLACK_FONT, LIGHT_GRAY, P)

R = 20  # Al revenue change
sc(ws3, R, 1, "Al Revenue Change", BLACK_FONT)
for j in range(8):
    sc(ws3, R, 2+j, f"={get_column_letter(2+j)}19*$B$12", BLACK_FONT, fmt=P)

R = 21  # Alumina revenue change
sc(ws3, R, 1, "Alumina Revenue Change", BLACK_FONT)
for j in range(8):
    sc(ws3, R, 2+j, f"={get_column_letter(2+j)}19*(1-$B$12)*$B$13", BLACK_FONT, fmt=P)

R = 22  # Total revenue change
sc(ws3, R, 1, "Total Revenue Change", SUBHEADER_FONT, YELLOW)
for j in range(8):
    cl = get_column_letter(2+j)
    sc(ws3, R, 2+j, f"={cl}20+{cl}21", RESULT_FONT, YELLOW, P)

R = 23  # Implied Revenue
sc(ws3, R, 1, "FY2025E Implied Revenue (M)", SUBHEADER_FONT, LIGHT_GRAY)
for j in range(8):
    cl = get_column_letter(2+j)
    sc(ws3, R, 2+j, f"=$B$9*(1+{cl}22)*(1+$B$15)", BLACK_FONT, LIGHT_GRAY, N)

R = 24  # Implied EBIT
sc(ws3, R, 1, "FY2025E Implied EBIT (M)", SUBHEADER_FONT)
for j in range(8):
    cl = get_column_letter(2+j)
    sc(ws3, R, 2+j, f"={cl}23*0.13*(1+{cl}19*$B$14*0.5)", BLACK_FONT, fmt=N)

R = 25  # EBIT Margin
sc(ws3, R, 1, "Implied EBIT Margin", BLACK_FONT)
for j in range(8):
    cl = get_column_letter(2+j)
    sc(ws3, R, 2+j, f"={cl}24/{cl}23", BLACK_FONT, fmt=P)

R = 26  # FCF
sc(ws3, R, 1, "Implied Unlevered FCF (M)", SUBHEADER_FONT)
for j in range(8):
    cl = get_column_letter(2+j)
    sc(ws3, R, 2+j, f"={cl}24*(1-DCF!$B$41)+{cl}23*DCF!$B$38-{cl}23*DCF!$B$39", BLACK_FONT, fmt=N)

R = 28
hdr(ws3, R, "IMPLIED SHARE PRICE BY AL PRICE", 1, 9)
R = 29
sc(ws3, R, 1, "SHFE Al Price (RMB/ton)", SUBHEADER_FONT, LIGHT_BLUE)
for j, p in enumerate(alps):
    sc(ws3, R, 2+j, p, SUBHEADER_FONT, LIGHT_BLUE, N)

R = 30  # Implied Price
sc(ws3, R, 1, "IMPLIED SHARE PRICE", RED_12, YELLOW)
for j in range(8):
    cl = get_column_letter(2+j)
    c = ws3.cell(row=R, column=2+j, value=f"=({cl}26/(DCF!$B$57-DCF!$B$37)+DCF!B70)/DCF!B72")
    c.font = RED_12; c.fill = YELLOW; c.number_format = PR; c.alignment = C; c.border = BORDER

R = 31  # Upside
sc(ws3, R, 1, "Upside vs Current Price", SUBHEADER_FONT, LIGHT_GRAY)
for j in range(8):
    cl = get_column_letter(2+j)
    sc(ws3, R, 2+j, f"={cl}30/DCF!B9-1", BLACK_FONT, LIGHT_GRAY, P)

# Dual-axis: Al Price x EBIT Margin
R = 34
hdr(ws3, R, "AL PRICE x EBIT MARGIN -> IMPLIED SHARE PRICE", 1, 9)
R = 35
ems = [0.09, 0.11, 0.12, 0.13, 0.14, 0.15]
als = [20000, 22000, 23000, 24000, 25000, 26000]
sc(ws3, R, 1, "AL \\ EBIT%", SUBHEADER_FONT, LIGHT_BLUE)
for j, e in enumerate(ems):
    sc(ws3, R, 2+j, e, SUBHEADER_FONT, LIGHT_BLUE, P)
for i, ap in enumerate(als):
    r = R+1+i
    sc(ws3, r, 1, ap, BLUE_FONT, fmt=N)
    for j, e in enumerate(ems):
        ac = f"$A{r}"
        ec = f"${get_column_letter(2+j)}${R}"
        rf = f"(1+({ac}/$B$8-1)*($B$12+(1-$B$12)*$B$13))*(1+$B$15)"
        rv = f"($B$9*{rf})"
        fc = f"({rv}*{ec}*(1-DCF!$B$41)+{rv}*DCF!$B$38-{rv}*DCF!$B$39)"
        sc(ws3, r, 2+j, f"=({fc}/(DCF!$B$57-DCF!$B$37)+DCF!B70)/DCF!B72", BLACK_FONT, fmt=PR)

# Notes
sc(ws3, 43, 1, "Notes:", SUBHEADER_FONT)
sc(ws3, 44, 1, "1. FY2024A avg realized Al price ~20,200 RMB/ton (estimated)", BLACK_FONT)
sc(ws3, 45, 1, "2. Al products ~65% of revenue, alumina & other ~35%", BLACK_FONT)
sc(ws3, 46, 1, "3. Chalco has higher operating leverage (1.8x) than Hongqiao due to more integrated ops", BLACK_FONT)
sc(ws3, 47, 1, "4. Alumina price ~55% correlated to aluminum price moves", BLACK_FONT)
sc(ws3, 48, 1, "5. Valuation uses Y1 FCF perpetuity method (simplified DCF)", BLACK_FONT)

# ============================================================
# SAVE
# ============================================================
out = r"c:\Users\rickylu\.gemini\antigravity\scratch\financial-services-plugins\601600_DCF_Model_2026-03-09.xlsx"
wb.save(out)
print(f"[OK] Saved: {out}")
print(f"Ticker: {TICKER}")
print(f"Price: {PRICE}")
print(f"MCap: {MCAP:,.0f}M")
print(f"WACC: {WACC:.2%}")
print(f"Shares: {SHARES:,}M")
