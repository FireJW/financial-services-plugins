# -*- coding: utf-8 -*-
'''
Envicool (002837.SZ) DCF Valuation Model
NVIDIA Liquid Cooling Core Supplier / Precision Thermal Management
Generated: 2026-03-16
'''

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

wb = openpyxl.Workbook()

BF = Font(name="Microsoft YaHei", size=11, color="0000FF")
KF = Font(name="Microsoft YaHei", size=11, color="000000")
GF = Font(name="Microsoft YaHei", size=11, color="008000")
HF = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
SF = Font(name="Microsoft YaHei", size=11, bold=True, color="000000")
TF = Font(name="Microsoft YaHei", size=14, bold=True, color="000000")
RF = Font(name="Microsoft YaHei", size=12, bold=True, color="000000")
RB = Font(name="Microsoft YaHei", size=14, bold=True, color="FF0000")

DB = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
LB = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
LG = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
YL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

CA = Alignment(horizontal="center", vertical="center")
BD = Border(left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"))
P='0.0%'; N='#,##0'; N1='#,##0.0'; PR='#,##0.00'

def sc(ws, r, c, v, f=None, fl=None, fm=None, cm=None):
    cell = ws.cell(row=r, column=c, value=v)
    if f: cell.font = f
    if fl: cell.fill = fl
    if fm: cell.number_format = fm
    cell.alignment = CA; cell.border = BD
    if cm: cell.comment = Comment(cm, "DCF")
    return cell

def hdr(ws, r, t, c1=1, c2=10):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c); cell.fill = DB; cell.font = HF; cell.alignment = CA; cell.border = BD
    ws.cell(row=r, column=c1, value=t)
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

def shdr(ws, r, t, c1=1, c2=10):
    for c in range(c1, c2+1):
        cell = ws.cell(row=r, column=c); cell.fill = LB; cell.font = SF; cell.alignment = CA; cell.border = BD
    ws.cell(row=r, column=c1, value=t)

TICKER = "002837.SZ"
DATE = "2026-03-16"
PRICE = 105.00
SHARES = 977
MCAP = PRICE * SHARES

hist = {
    "year":    [2021,    2022,     2023,     2024],
    "rev":     [2228,    2923,     3529,     4589],
    "ebit":    [240,     350,      440,      590],
    "ebit_pct":[0.108,   0.120,    0.125,    0.129],
    "tax":     [30,      45,       55,       75],
    "ni":      [205,     280,      344,      453],
    "da":      [65,      80,       100,      151],
    "capex":   [80,      120,      204,      280],
    "ebitda":  [305,     430,      540,      741],
}

DEBT = 1400
CASH = 600
NET_DEBT = DEBT - CASH

Rf = 0.022
BETA = 0.27
ERP = 0.065
KE = Rf + BETA * ERP
KD_PRE = 0.038
TAX_R = 0.15
KD_POST = KD_PRE * (1 - TAX_R)
EV = MCAP + NET_DEBT
EQ_W = MCAP / EV
DT_W = NET_DEBT / EV
WACC_V = KE * EQ_W + KD_POST * DT_W

# High-growth AI liquid cooling company - aggressive projections justified
proj = {
    "years": [2025, 2026, 2027, 2028, 2029],
    "bear":  {"rg": [0.20, 0.15, 0.12, 0.10, 0.08], "em": [0.11, 0.11, 0.11, 0.10, 0.10]},
    "base":  {"rg": [0.35, 0.30, 0.25, 0.18, 0.12], "em": [0.13, 0.14, 0.14, 0.13, 0.13]},
    "bull":  {"rg": [0.50, 0.40, 0.30, 0.22, 0.15], "em": [0.15, 0.16, 0.17, 0.16, 0.15]},
}
DA_P = 0.030
CX_P = 0.055
NW_P = 0.020
TG = {"bear": 0.025, "base": 0.035, "bull": 0.045}

ws = wb.active; ws.title = "DCF"; ws.sheet_properties.tabColor = "17365D"
ws.column_dimensions['A'].width = 32
for c in range(2, 12): ws.column_dimensions[get_column_letter(c)].width = 16

sc(ws, 1, 1, f"Envicool (002837.SZ) DCF Model", TF)
ws.merge_cells('A1:J1')
sc(ws, 2, 1, f"TICKER: {TICKER} | DATE: {DATE} | FY2024", KF)
ws.merge_cells('A2:J2')
sc(ws, 3, 1, "Unit: RMB Millions (M) | NVIDIA Liquid Cooling Core Supplier", KF)
ws.merge_cells('A3:J3')

hdr(ws, 5, "CASE SELECTOR", 1, 10)
sc(ws, 6, 1, "Active Case (1=Bear, 2=Base, 3=Bull)", SF, YL)
sc(ws, 6, 2, 2, BF, YL, cm="User input: 1/2/3")

hdr(ws, 8, "MARKET DATA", 1, 10)
mkt = [("Stock Price", PRICE, PR, "Source: 2026-03-16"),
       ("Shares Outstanding (M)", SHARES, N, "Source: 9.77 yi"),
       ("Market Cap (M)", None, N, None),
       ("Total Debt (M)", DEBT, N, "Est. interest-bearing debt"),
       ("Cash (M)", CASH, N, "Source: FY2024"),
       ("Net Debt (M)", None, N, None),
       ("Enterprise Value (M)", None, N, None)]
for i, (lb, v, fm, cm) in enumerate(mkt):
    r = 9 + i
    sc(ws, r, 1, lb, SF, GY)
    if v is not None: sc(ws, r, 2, v, BF, LG, fm, cm)
    elif i == 2: sc(ws, r, 2, "=B9*B10", KF, fm=fm)
    elif i == 5: sc(ws, r, 2, "=B12-B13", KF, fm=fm)
    elif i == 6: sc(ws, r, 2, "=B11+B14", KF, fm=fm)

hdr(ws, 17, "SCENARIO ASSUMPTIONS", 1, 10)
hp = ["Assumption"] + [f"FY{y}E" for y in proj["years"]]

def wc(ws, sr, name, ck, tgk):
    shdr(ws, sr, name, 1, 10)
    for c, h in enumerate(hp, 1): sc(ws, sr+1, c, h, SF, LB)
    r = sr + 2
    sc(ws, r, 1, "Revenue Growth", SF, GY)
    for j, v in enumerate(proj[ck]["rg"]): sc(ws, r, 2+j, v, BF, LG, P)
    r += 1
    sc(ws, r, 1, "EBIT Margin", SF, GY)
    for j, v in enumerate(proj[ck]["em"]): sc(ws, r, 2+j, v, BF, LG, P)
    r += 1
    sc(ws, r, 1, "Terminal Growth", SF, GY)
    sc(ws, r, 2, TG[tgk], BF, LG, P)

wc(ws, 18, "BEAR CASE", "bear", "bear")
wc(ws, 23, "BASE CASE", "base", "base")
wc(ws, 28, "BULL CASE", "bull", "bull")

shdr(ws, 33, "SELECTED CASE (B6)", 1, 10)
for c, h in enumerate(hp, 1): sc(ws, 34, c, h, SF, LB)
sc(ws, 35, 1, "Revenue Growth", SF, YL)
for j in range(5):
    cl = get_column_letter(2+j)
    sc(ws, 35, 2+j, f'=IF($B$6=1,{cl}$20,IF($B$6=2,{cl}$25,{cl}$30))', KF, YL, P)
sc(ws, 36, 1, "EBIT Margin", SF, YL)
for j in range(5):
    cl = get_column_letter(2+j)
    sc(ws, 36, 2+j, f'=IF($B$6=1,{cl}$21,IF($B$6=2,{cl}$26,{cl}$31))', KF, YL, P)
sc(ws, 37, 1, "Terminal Growth", SF, YL)
sc(ws, 37, 2, '=IF($B$6=1,$B$22,IF($B$6=2,$B$27,$B$32))', KF, YL, P)
sc(ws, 38, 1, "D&A % Rev", SF, GY); sc(ws, 38, 2, DA_P, BF, LG, P, "~3% asset-light model")
sc(ws, 39, 1, "CapEx % Rev", SF, GY); sc(ws, 39, 2, CX_P, BF, LG, P, "~5.5% incl Malaysia plant")
sc(ws, 40, 1, "NWC % dRev", SF, GY); sc(ws, 40, 2, NW_P, BF, LG, P)
sc(ws, 41, 1, "Tax Rate", SF, GY); sc(ws, 41, 2, TAX_R, BF, LG, P, "15% HNTE preferential")

hdr(ws, 43, "INCOME STATEMENT & FREE CASH FLOW", 1, 10)
yrs = [""] + [f"FY{y}A" for y in hist["year"]] + [f"FY{y}E" for y in proj["years"]]
for c, h in enumerate(yrs, 1):
    fl = LB if "A" in str(h) else GY if "E" in str(h) else None
    sc(ws, 44, c, h, SF, fl)

sc(ws, 45, 1, "Revenue", SF)
for i, rv in enumerate(hist["rev"]): sc(ws, 45, 2+i, rv, BF, fm=N, cm=f"FY{hist['year'][i]}")
for j in range(5): sc(ws, 45, 6+j, f"={get_column_letter(5+j)}45*(1+{get_column_letter(2+j)}$35)", KF, fm=N)

sc(ws, 46, 1, "  Growth %", KF)
for i in range(1, 4): sc(ws, 46, 2+i, f"={get_column_letter(2+i)}45/{get_column_letter(1+i)}45-1", KF, fm=P)
for j in range(5): sc(ws, 46, 6+j, f"={get_column_letter(6+j)}45/{get_column_letter(5+j)}45-1", KF, fm=P)

sc(ws, 47, 1, "EBIT", SF)
for i, eb in enumerate(hist["ebit"]): sc(ws, 47, 2+i, eb, BF, fm=N)
for j in range(5): sc(ws, 47, 6+j, f"={get_column_letter(6+j)}45*{get_column_letter(2+j)}$36", KF, fm=N)

sc(ws, 48, 1, "  EBIT Margin", KF)
for i in range(4): sc(ws, 48, 2+i, f"={get_column_letter(2+i)}47/{get_column_letter(2+i)}45", KF, fm=P)
for j in range(5): sc(ws, 48, 6+j, f"={get_column_letter(6+j)}47/{get_column_letter(6+j)}45", KF, fm=P)

sc(ws, 49, 1, "(-) Taxes", KF)
for i, tx in enumerate(hist["tax"]): sc(ws, 49, 2+i, -abs(tx), BF, fm=N)
for j in range(5): sc(ws, 49, 6+j, f"=-{get_column_letter(6+j)}47*$B$41", KF, fm=N)

sc(ws, 50, 1, "NOPAT", SF)
for i in range(4): sc(ws, 50, 2+i, f"={get_column_letter(2+i)}47+{get_column_letter(2+i)}49", KF, fm=N)
for j in range(5): sc(ws, 50, 6+j, f"={get_column_letter(6+j)}47+{get_column_letter(6+j)}49", KF, fm=N)

sc(ws, 51, 1, "(+) D&A", KF)
for i, da in enumerate(hist["da"]): sc(ws, 51, 2+i, da, BF, fm=N)
for j in range(5): sc(ws, 51, 6+j, f"={get_column_letter(6+j)}45*$B$38", KF, fm=N)

sc(ws, 52, 1, "(-) CapEx", KF)
for i, cx in enumerate(hist["capex"]): sc(ws, 52, 2+i, -abs(cx), BF, fm=N)
for j in range(5): sc(ws, 52, 6+j, f"=-{get_column_letter(6+j)}45*$B$39", KF, fm=N)

sc(ws, 53, 1, "(-) dNWC", KF)
for i in range(4): sc(ws, 53, 2+i, 0, BF, fm=N)
for j in range(5): sc(ws, 53, 6+j, f"=-({get_column_letter(6+j)}45-{get_column_letter(5+j)}45)*$B$40", KF, fm=N)

sc(ws, 54, 1, "Unlevered FCF", SF, GY)
for i in range(4):
    cl = get_column_letter(2+i)
    sc(ws, 54, 2+i, f"={cl}50+{cl}51+{cl}52+{cl}53", RF, GY, N)
for j in range(5):
    cl = get_column_letter(6+j)
    sc(ws, 54, 6+j, f"={cl}50+{cl}51+{cl}52+{cl}53", RF, GY, N)

hdr(ws, 56, "DCF VALUATION", 1, 10)
sc(ws, 57, 1, "WACC", SF, YL); sc(ws, 57, 2, "=WACC!B16", GF, YL, P)
sc(ws, 58, 1, "Discount Period", SF)
for j in range(5): sc(ws, 58, 6+j, 0.5+j, BF, fm=N1)
sc(ws, 59, 1, "Discount Factor", KF)
for j in range(5): sc(ws, 59, 6+j, f"=1/(1+$B$57)^{get_column_letter(6+j)}58", KF, fm='0.0000')
sc(ws, 60, 1, "PV of FCF", SF)
for j in range(5):
    cl = get_column_letter(6+j)
    sc(ws, 60, 6+j, f"={cl}54*{cl}59", KF, fm=N)

sc(ws, 62, 1, "Terminal FCF", KF); sc(ws, 62, 2, "=J54*(1+B37)", KF, fm=N)
sc(ws, 63, 1, "Terminal Value", KF); sc(ws, 63, 2, "=B62/(B57-B37)", KF, fm=N)
sc(ws, 64, 1, "PV of Terminal Value", SF); sc(ws, 64, 2, "=B63/(1+B57)^4.5", KF, fm=N)

hdr(ws, 66, "VALUATION SUMMARY", 1, 10)
sc(ws, 67, 1, "Sum PV of FCFs", SF, GY); sc(ws, 67, 2, "=SUM(F60:J60)", KF, GY, N)
sc(ws, 68, 1, "PV of Terminal Value", SF, GY); sc(ws, 68, 2, "=B64", KF, GY, N)
sc(ws, 69, 1, "Enterprise Value", SF, YL); sc(ws, 69, 2, "=B67+B68", RF, YL, N)
sc(ws, 70, 1, "(-) Net Debt", SF, GY); sc(ws, 70, 2, "=-B14", KF, GY, N)
sc(ws, 71, 1, "Equity Value", SF, YL); sc(ws, 71, 2, "=B69+B70", RF, YL, N)
sc(ws, 72, 1, "Shares (M)", KF); sc(ws, 72, 2, "=B10", GF, fm=N)
sc(ws, 73, 1, "IMPLIED PRICE", RB, YL); sc(ws, 73, 2, "=B71/B72", RB, YL, PR)
sc(ws, 74, 1, "Current Price", SF); sc(ws, 74, 2, "=B9", GF, fm=PR)
sc(ws, 75, 1, "Implied Upside/(Downside)", SF, YL); sc(ws, 75, 2, "=B73/B74-1", RF, YL, P)

# Sensitivity 1: WACC vs TGR
hdr(ws, 77, "SENSITIVITY ANALYSIS", 1, 10)
shdr(ws, 78, "Table 1: Price vs WACC & Terminal Growth", 1, 8)
tgv = [0.020, 0.025, 0.030, 0.035, 0.045]
wv = [0.03, 0.04, 0.05, 0.06, 0.07]
sc(ws, 79, 1, "WACC \\ TGR", SF, LB)
for j, t in enumerate(tgv): sc(ws, 79, 2+j, t, SF, LB, P)
for i, w in enumerate(wv):
    r = 80+i
    sc(ws, r, 1, w, BF, fm=P)
    for j, t in enumerate(tgv):
        pvs = "+".join([f"{get_column_letter(6+k)}54/(1+$A{r})^{0.5+k}" for k in range(5)])
        tc = f"${get_column_letter(2+j)}$79"; wcc = f"$A{r}"
        tv = f"(J54*(1+{tc})/({wcc}-{tc}))/(1+{wcc})^4.5"
        sc(ws, r, 2+j, f"=({pvs}+{tv}+B70)/B72", KF, fm=PR)

shdr(ws, 86, "Table 2: Price vs Rev Growth & EBIT Margin", 1, 8)
rgv = [0.15, 0.20, 0.25, 0.30, 0.40]
emv = [0.10, 0.12, 0.14, 0.16, 0.18]
sc(ws, 87, 1, "RevGr \\ EBIT%", SF, LB)
for j, e in enumerate(emv): sc(ws, 87, 2+j, e, SF, LB, P)
for i, rg in enumerate(rgv):
    r = 88+i
    sc(ws, r, 1, rg, BF, fm=P)
    for j, e in enumerate(emv):
        ec = f"${get_column_letter(2+j)}$87"; rc = f"$A{r}"
        fcf = f"(E45*(1+{rc})*{ec}*(1-$B$41)+E45*(1+{rc})*$B$38-E45*(1+{rc})*$B$39-E45*{rc}*$B$40)"
        ag = f"({rc}/2+$B$37/2)"
        sc(ws, r, 2+j, f"=({fcf}/($B$57-{ag})+B70)/B72", KF, fm=PR)

shdr(ws, 94, "Table 3: Price vs Beta & Risk-Free Rate", 1, 8)
bv = [0.15, 0.20, 0.27, 0.35, 0.50]
rv = [0.015, 0.020, 0.022, 0.025, 0.030]
sc(ws, 95, 1, "Beta \\ Rf", SF, LB)
for j, rf in enumerate(rv): sc(ws, 95, 2+j, rf, SF, LB, P)
for i, b in enumerate(bv):
    r = 96+i
    sc(ws, r, 1, b, BF, fm='0.00')
    for j, rf in enumerate(rv):
        rc = f"${get_column_letter(2+j)}$95"; bc = f"$A{r}"
        ke = f"({rc}+{bc}*{ERP})"
        wn = f"({ke}*{EQ_W:.4f}+{KD_POST:.4f}*{DT_W:.4f})"
        pvs = "+".join([f"{get_column_letter(6+k)}54/(1+{wn})^{0.5+k}" for k in range(5)])
        tv = f"(J54*(1+B37)/({wn}-B37))/(1+{wn})^4.5"
        sc(ws, r, 2+j, f"=({pvs}+{tv}+B70)/B72", KF, fm=PR)

# WACC SHEET
ws2 = wb.create_sheet("WACC"); ws2.sheet_properties.tabColor = "4472C4"
ws2.column_dimensions['A'].width = 40; ws2.column_dimensions['B'].width = 18
hdr(ws2, 1, "WACC CALCULATION", 1, 3)
shdr(ws2, 2, "COST OF EQUITY (CAPM)", 1, 3)
sc(ws2, 3, 1, "Risk-Free Rate (10Y)", KF); sc(ws2, 3, 2, Rf, BF, LG, P, "China 10Y bond")
sc(ws2, 4, 1, "Beta (5Y monthly vs CSI300)", KF); sc(ws2, 4, 2, BETA, BF, LG, '0.00', "Very low beta - defensive thermal mgmt")
sc(ws2, 5, 1, "Equity Risk Premium", KF); sc(ws2, 5, 2, ERP, BF, LG, P)
sc(ws2, 6, 1, "Cost of Equity", SF, GY); sc(ws2, 6, 2, "=B3+B4*B5", RF, GY, P)
shdr(ws2, 8, "COST OF DEBT", 1, 3)
sc(ws2, 9, 1, "Pre-Tax Cost of Debt", KF); sc(ws2, 9, 2, KD_PRE, BF, LG, P)
sc(ws2, 10, 1, "Tax Rate", KF); sc(ws2, 10, 2, TAX_R, BF, LG, P, "15% HNTE")
sc(ws2, 11, 1, "After-Tax Cost of Debt", SF, GY); sc(ws2, 11, 2, "=B9*(1-B10)", RF, GY, P)
shdr(ws2, 13, "CAPITAL STRUCTURE", 1, 3)
sc(ws2, 14, 1, "Equity Weight", KF); sc(ws2, 14, 2, "=DCF!B11/(DCF!B11+DCF!B14)", GF, fm=P)
sc(ws2, 15, 1, "Debt Weight", KF); sc(ws2, 15, 2, "=DCF!B14/(DCF!B11+DCF!B14)", GF, fm=P)
sc(ws2, 16, 1, "WACC", RB, YL); sc(ws2, 16, 2, "=B6*B14+B11*B15", RB, YL, P)

# SEGMENTS SHEET
ws3 = wb.create_sheet("Segments"); ws3.sheet_properties.tabColor = "548235"
ws3.column_dimensions['A'].width = 38
for c in range(2, 6): ws3.column_dimensions[get_column_letter(c)].width = 16
hdr(ws3, 1, "BUSINESS SEGMENT ANALYSIS (FY2024)", 1, 5)
shdr(ws3, 2, "Revenue by Segment (M RMB)", 1, 5)
sc(ws3, 3, 1, "Segment", SF, LB); sc(ws3, 3, 2, "Revenue", SF, LB)
sc(ws3, 3, 3, "% of Total", SF, LB); sc(ws3, 3, 4, "Key Driver", SF, LB)
segs = [
    ("DC Thermal (incl Liquid Cooling)", 2750, 0.599, "NVIDIA GB300 orders, AI DC boom"),
    ("Telecom Network Thermal", 950, 0.207, "5G base station, edge computing"),
    ("Rail Transit Thermal", 500, 0.109, "Metro/HSR HVAC systems"),
    ("Other (EV, Industrial)", 389, 0.085, "EV thermal, new segments"),
]
for i, (nm, rv, pct, dr) in enumerate(segs):
    r = 4 + i
    sc(ws3, r, 1, nm, SF, GY); sc(ws3, r, 2, rv, BF, fm=N, cm="FY2024 est")
    sc(ws3, r, 3, pct, KF, fm=P); sc(ws3, r, 4, dr, KF)
sc(ws3, 8, 1, "Total Revenue", SF, YL); sc(ws3, 8, 2, "=SUM(B4:B7)", RF, YL, N)
sc(ws3, 8, 3, "=SUM(C4:C7)", RF, YL, P)
sc(ws3, 10, 1, "KEY NOTES:", SF)
sc(ws3, 11, 1, "1. NVIDIA core supplier: UQD quick-connects (global 1st MGX-certified)", KF)
sc(ws3, 12, 1, "2. GB300 value share 35% per cabinet (highest China supplier)", KF)
sc(ws3, 13, 1, "3. 1200-unit GB300 liquid cooling cabinet order = RMB 9.6-10.2B", KF)
sc(ws3, 14, 1, "4. Malaysia liquid cooling plant commissioning Q4 2025", KF)
sc(ws3, 15, 1, "5. Liquid cooling rev ~2x YoY in FY2024, 30% of DC business", KF)
sc(ws3, 16, 1, "6. COOLERCHIPS strategic tech partner with NVIDIA", KF)

out = r"c:\Users\rickylu\.gemini\antigravity\scratch\financial-services-plugins\002837_DCF_Model_2026-03-16.xlsx"
wb.save(out)
print(f"[OK] Saved: {out}")
print(f"Ticker: {TICKER}")
print(f"Price: {PRICE}")
print(f"MCap: {MCAP:,.0f}M")
print(f"WACC: {WACC_V:.2%}")
